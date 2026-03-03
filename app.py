#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PROCESSADOR RV E DSR - BACKEND WEB
Sistema com autenticação e gerenciamento de usuários/acionistas
"""

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, redirect, url_for, session
from functools import wraps
import pdfplumber
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
import json
import hashlib
import zipfile
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'output_rv'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
app.secret_key = 'totvs-nordeste-conciliacao-2024-secret-key'

# Cria pastas se não existirem
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# ==================================================================================
# ARQUIVOS DE DADOS
# ==================================================================================

USUARIOS_FILE = 'data/usuarios.json'
ACIONISTAS_FILE = 'data/acionistas.json'
ROYALTIES_CONFIG_FILE = 'data/royalties_config.json'
ROYALTIES_OUTPUT_FOLDER = 'output_royalties'
CONCILIACAO_OUTPUT_FOLDER = 'output_conciliacao'

NOTAS_OUTPUT_FOLDER = 'output_notas'

os.makedirs('data', exist_ok=True)
os.makedirs(ROYALTIES_OUTPUT_FOLDER, exist_ok=True)
os.makedirs(CONCILIACAO_OUTPUT_FOLDER, exist_ok=True)
os.makedirs(NOTAS_OUTPUT_FOLDER, exist_ok=True)

# Import do motor de conciliação bancária x contábil
try:
    from conciliacao_bancaria_contabil import processar_conciliacao
except ImportError:
    processar_conciliacao = None

# Import do motor de conciliação cliente (NFs x Recebimentos)
try:
    from conciliacao_bancaria_cliente import processar_conciliacao_cliente
except ImportError:
    processar_conciliacao_cliente = None

def hash_password(password):
    """Hash simples para senhas"""
    return hashlib.sha256(password.encode()).hexdigest()

def load_usuarios():
    """Carrega usuários do arquivo JSON"""
    if os.path.exists(USUARIOS_FILE):
        with open(USUARIOS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    # Usuário admin padrão
    default_users = {
        'admin': {
            'name': 'Administrador',
            'password': hash_password('admin123'),
            'is_admin': True,
            'active': True
        }
    }
    save_usuarios(default_users)
    return default_users

def save_usuarios(usuarios):
    """Salva usuários no arquivo JSON"""
    with open(USUARIOS_FILE, 'w', encoding='utf-8') as f:
        json.dump(usuarios, f, ensure_ascii=False, indent=2)

def load_acionistas():
    """Carrega acionistas do arquivo JSON"""
    if os.path.exists(ACIONISTAS_FILE):
        with open(ACIONISTAS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    # Acionistas padrão
    default_acionistas = {
        "Francisco Ferreira": {"active": True},
        "Marcos Saraiva": {"active": True},
        "Julio Bernadotte": {"active": True},
        "Luciana Colares": {"active": True},
        "Marcos Guimaraes": {"active": True},
        "Maury Neto": {"active": True},
        "Paulo Morais": {"active": True},
        "Romulo Barroso": {"active": True},
        "Simone Guimaraes": {"active": True},
    }
    save_acionistas(default_acionistas)
    return default_acionistas

def save_acionistas(acionistas):
    """Salva acionistas no arquivo JSON"""
    with open(ACIONISTAS_FILE, 'w', encoding='utf-8') as f:
        json.dump(acionistas, f, ensure_ascii=False, indent=2)


def load_royalties_config():
    """Carrega configuração de royalties do arquivo JSON"""
    default_config = {
        "produtos_nao_royalties": [
            "PS01009", "PS01010", "PS02001", "PS02002", "PS02003",
            "PS02004", "PS02010", "PS02020", "PS02021", "PS02022",
            "PS02023", "PS02030", "PS02031", "PS03001"
        ],
        "clientes_nao_royalties": [
            "DEBITO", "A00001", "A00002", "A000AL", "A00158", "A84063",
            "AAA002", "AAA003", "AAA004", "AAA005", "AAA006", "AAA007",
            "AAA008", "T82665", "TEZBY4", "TFDBWH", "TFCNN8", "TFCPGR",
            "TFEBT1", "TFEFYX", "X000LI", "X00111", "X00112", "X0004I",
            "X0010W", "X0012G", "X000LA", "X000EE", "TFDVD4", "TFDHQX",
            "TEZGHH", "A000BY"
        ]
    }

    if os.path.exists(ROYALTIES_CONFIG_FILE):
        try:
            with open(ROYALTIES_CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            save_royalties_config(default_config)
            return default_config
    else:
        save_royalties_config(default_config)
        return default_config


def save_royalties_config(config):
    """Salva configuração de royalties no arquivo JSON"""
    with open(ROYALTIES_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

# ==================================================================================
# DECORADORES DE AUTENTICAÇÃO
# ==================================================================================

def login_required(f):
    """Decorator para rotas que requerem login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator para rotas que requerem admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        if not session.get('is_admin'):
            return jsonify({'error': 'Acesso negado'}), 403
        return f(*args, **kwargs)
    return decorated_function

# ==================================================================================
# MAPEAMENTO DE ACIONISTAS (agora dinâmico via JSON)
# ==================================================================================

def get_acionistas_set():
    """Retorna set de acionistas ativos"""
    acionistas = load_acionistas()
    return {nome for nome, data in acionistas.items() if data.get('active', True)}

# ==================================================================================
# FUNÇÕES DE PROCESSAMENTO
# ==================================================================================

def eh_acionista(nome):
    """Verifica se o funcionário é ACIONISTA"""
    acionistas = get_acionistas_set()
    if nome in acionistas:
        return True

    nome_limpo = nome.upper().strip()
    for acionista in acionistas:
        if acionista.upper() == nome_limpo:
            return True

    return False

def pegar_valor(valor):
    """Converte valor para float"""
    if not valor or valor == '-':
        return 0.0
    
    if isinstance(valor, (int, float)):
        return float(valor)
    
    valor = str(valor).replace('R$', '').replace(' ', '').strip()
    
    if ',' in valor and '.' in valor:
        valor = valor.replace('.', '').replace(',', '.')
    elif ',' in valor:
        valor = valor.replace(',', '.')
    
    try:
        return float(valor)
    except:
        return 0.0

def formatar(valor):
    """Formata valor com 18 caracteres: 15 dígitos + ponto + 2 centavos"""
    valor_float = pegar_valor(valor)
    centavos = int(round(valor_float * 100))
    parte_inteira = centavos // 100
    parte_decimal = centavos % 100
    return f"{parte_inteira:015d}.{parte_decimal:02d}"

def processar_arquivo_rv_dsr(arquivo_path):
    """Processa o arquivo XLSX e retorna estatísticas"""
    
    # Abre arquivo com data_only=True para obter valores calculados
    wb = openpyxl.load_workbook(arquivo_path, data_only=True)

    # Tenta abrir a aba "MODELO PARA AUTOMAÇÃO"
    aba_nome = "MODELO PARA AUTOMAÇÃO"
    if aba_nome not in wb.sheetnames:
        for nome in wb.sheetnames:
            if "MODELO" in nome.upper() and "AUTOMA" in nome.upper():
                aba_nome = nome
                break
        else:
            raise Exception(f"Aba 'MODELO PARA AUTOMAÇÃO' não encontrada!\nAbas disponíveis: {', '.join(wb.sheetnames)}")

    ws = wb[aba_nome]

    # Processa linhas
    dados_por_filial = {}
    total = 0
    acionistas_count = 0
    clt_count = 0
    ignorados_filial = 0
    ignorados_valor = 0

    log_processados = []
    log_ignorados = []

    for i in range(2, ws.max_row + 1):
        executivo = ws.cell(i, 3).value  # Coluna C
        rv_raw = ws.cell(i, 8).value     # Coluna H
        dsr_raw = ws.cell(i, 9).value    # Coluna I
        filial = ws.cell(i, 12).value    # Coluna L
        matricula = ws.cell(i, 13).value # Coluna M

        if not executivo:
            continue

        executivo = str(executivo).strip()

        # Valida filial/matrícula
        if not filial or not matricula or str(filial).upper() == "VAZIO" or str(matricula).upper() == "VAZIO":
            log_ignorados.append({
                "executivo": executivo,
                "motivo": "Sem filial/matrícula",
                "rv": rv_raw,
                "dsr": dsr_raw
            })
            ignorados_filial += 1
            continue

        filial = str(filial).strip().zfill(6)         # filial sempre 6
        matricula = str(matricula).strip().zfill(6)   # matrícula sempre 6 (padrão layout 2)

        rv_valor = pegar_valor(rv_raw)
        dsr_valor = pegar_valor(dsr_raw)

        # Valida valores
        if rv_valor == 0 and dsr_valor == 0:
            log_ignorados.append({
                "executivo": executivo,
                "motivo": "RV e DSR zerados/vazios",
                "rv": rv_raw,
                "dsr": dsr_raw
            })
            ignorados_valor += 1
            continue

        is_acionista = eh_acionista(executivo)

        if is_acionista:
            cod_rv = "390"
            cod_dsr = "391"
            tipo = "ACIONISTA"
            acionistas_count += 1
        else:
            cod_rv = "392"
            cod_dsr = "393"
            tipo = "CLT"
            clt_count += 1

        rv_fmt = formatar(rv_valor)
        dsr_fmt = formatar(dsr_valor)
        identificador = filial + matricula            # identificador sempre 12 chars
        espacos = 6                                   # espaçamento fixo (padrão layout 2)

        linha_rv = f"{identificador}{' ' * espacos}{cod_rv}{rv_fmt}"
        linha_dsr = f"{identificador}{' ' * espacos}{cod_dsr}{dsr_fmt}"

        if filial not in dados_por_filial:
            dados_por_filial[filial] = []

        dados_por_filial[filial].append(linha_rv)
        dados_por_filial[filial].append(linha_dsr)

        log_processados.append({
            "executivo": executivo,
            "tipo": tipo,
            "rv": rv_valor,
            "dsr": dsr_valor,
            "filial": filial,
            "matricula": matricula
        })

        total += 1

    # Gera arquivos TXT
    pasta = app.config['OUTPUT_FOLDER']
    arquivos_gerados = []

    for filial, linhas in dados_por_filial.items():
        arquivo = os.path.join(pasta, f"filial_{filial}.txt")
        with open(arquivo, 'w', encoding='utf-8') as f:
            for linha in linhas:
                f.write(linha + '\n')
        arquivos_gerados.append(f"filial_{filial}.txt")

    # Gera log detalhado
    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo_log = os.path.join(pasta, f"log_processamento_{data_hora}.txt")

    with open(arquivo_log, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("LOG DE PROCESSAMENTO RV/DSR\n")
        f.write(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write("="*80 + "\n\n")

        f.write("-"*80 + "\n")
        f.write(f"FUNCIONÁRIOS PROCESSADOS ({total})\n")
        f.write("-"*80 + "\n")
        for item in log_processados:
            f.write(f"  {item['executivo']:30} | {item['tipo']:10} | "
                   f"Filial: {item['filial']} | Mat: {item['matricula']} | "
                   f"RV: R$ {item['rv']:>10.2f} | DSR: R$ {item['dsr']:>10.2f}\n")
        f.write("\n")

        total_ignorados = ignorados_filial + ignorados_valor
        f.write("-"*80 + "\n")
        f.write(f"FUNCIONÁRIOS IGNORADOS ({total_ignorados})\n")
        f.write("-"*80 + "\n")
        if log_ignorados:
            for item in log_ignorados:
                f.write(f"  {item['executivo']:30} | Motivo: {item['motivo']}\n")
        else:
            f.write("  Nenhum funcionário ignorado.\n")
        f.write("\n")

        f.write("="*80 + "\n")
        f.write("RESUMO\n")
        f.write("="*80 + "\n")
        f.write(f"  Total processados: {total}\n")
        f.write(f"  - ACIONISTAS (390/391): {acionistas_count}\n")
        f.write(f"  - CLT (392/393): {clt_count}\n")
        f.write(f"  Total ignorados: {total_ignorados}\n")
        f.write(f"  Arquivos gerados: {len(dados_por_filial)}\n")
        f.write("="*80 + "\n")

    arquivos_gerados.append(f"log_processamento_{data_hora}.txt")

    return {
        'total': total,
        'acionistas': acionistas_count,
        'clt': clt_count,
        'ignorados': ignorados_filial + ignorados_valor,
        'ignorados_filial': ignorados_filial,
        'ignorados_valor': ignorados_valor,
        'filiais': len(dados_por_filial),
        'arquivos': arquivos_gerados,
        'log_arquivo': f"log_processamento_{data_hora}.txt"
    }

# ==================================================================================
# ROTAS DE AUTENTICAÇÃO
# ==================================================================================

@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/assets/<path:filename>')
def serve_assets(filename):
    return send_from_directory('assets', filename)

@app.route('/login')
def login():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')

    usuarios = load_usuarios()

    if username not in usuarios:
        return jsonify({'success': False, 'error': 'Usuário não encontrado'})

    user = usuarios[username]

    if not user.get('active', True):
        return jsonify({'success': False, 'error': 'Usuário inativo'})

    if user['password'] != hash_password(password):
        return jsonify({'success': False, 'error': 'Senha incorreta'})

    session['user'] = username
    session['name'] = user.get('name', username)
    session['is_admin'] = user.get('is_admin', False)

    return jsonify({'success': True, 'redirect': '/dashboard'})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })

@app.route('/rv-dsr')
@login_required
def rv_dsr():
    return render_template('rv_dsr.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })


@app.route('/royalties')
@login_required
def royalties():
    return render_template('royalties.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })

# ==================================================================================
# ROTAS DE ADMIN - USUÁRIOS
# ==================================================================================

@app.route('/admin/usuarios')
@admin_required
def admin_usuarios():
    return render_template('usuarios.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })

@app.route('/api/usuarios', methods=['GET'])
@admin_required
def api_get_usuarios():
    usuarios = load_usuarios()
    # Remove senhas da resposta
    safe_usuarios = {}
    for username, data in usuarios.items():
        safe_usuarios[username] = {k: v for k, v in data.items() if k != 'password'}
    return jsonify(safe_usuarios)

@app.route('/api/usuarios', methods=['POST'])
@admin_required
def api_create_usuario():
    data = request.get_json()
    username = data.get('username', '').strip().lower()
    name = data.get('name', '').strip()
    password = data.get('password', '') or 'mudar123'
    is_admin = data.get('is_admin', False)
    active = data.get('active', True)

    if not username:
        return jsonify({'success': False, 'error': 'Username é obrigatório'})

    usuarios = load_usuarios()

    if username in usuarios:
        return jsonify({'success': False, 'error': 'Usuário já existe'})

    usuarios[username] = {
        'name': name,
        'password': hash_password(password),
        'is_admin': is_admin,
        'active': active
    }

    save_usuarios(usuarios)
    return jsonify({'success': True, 'message': 'Usuário criado com sucesso!'})

@app.route('/api/usuarios/<username>', methods=['PUT'])
@admin_required
def api_update_usuario(username):
    data = request.get_json()
    usuarios = load_usuarios()

    if username not in usuarios:
        return jsonify({'success': False, 'error': 'Usuário não encontrado'})

    if 'name' in data:
        usuarios[username]['name'] = data['name']
    if 'is_admin' in data:
        usuarios[username]['is_admin'] = data['is_admin']
    if 'active' in data:
        usuarios[username]['active'] = data['active']
    if data.get('password'):
        usuarios[username]['password'] = hash_password(data['password'])

    save_usuarios(usuarios)
    return jsonify({'success': True, 'message': 'Usuário atualizado!'})

@app.route('/api/usuarios/<username>', methods=['DELETE'])
@admin_required
def api_delete_usuario(username):
    if username == 'admin':
        return jsonify({'success': False, 'error': 'Não é possível excluir o admin'})

    usuarios = load_usuarios()

    if username not in usuarios:
        return jsonify({'success': False, 'error': 'Usuário não encontrado'})

    del usuarios[username]
    save_usuarios(usuarios)
    return jsonify({'success': True, 'message': 'Usuário excluído!'})

# ==================================================================================
# ROTAS DE ADMIN - ACIONISTAS
# ==================================================================================

@app.route('/admin/acionistas')
@admin_required
def admin_acionistas():
    return render_template('acionistas.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })

@app.route('/api/acionistas', methods=['GET'])
@login_required
def api_get_acionistas():
    acionistas = load_acionistas()
    return jsonify(acionistas)

@app.route('/api/acionistas', methods=['POST'])
@admin_required
def api_create_acionista():
    data = request.get_json()
    nome = data.get('nome', '').strip()
    active = data.get('active', True)

    if not nome:
        return jsonify({'success': False, 'error': 'Nome é obrigatório'})

    acionistas = load_acionistas()

    if nome in acionistas:
        return jsonify({'success': False, 'error': 'Acionista já existe'})

    acionistas[nome] = {'active': active}
    save_acionistas(acionistas)
    return jsonify({'success': True, 'message': 'Acionista cadastrado!'})

@app.route('/api/acionistas/<nome>', methods=['PUT'])
@admin_required
def api_update_acionista(nome):
    data = request.get_json()
    acionistas = load_acionistas()

    if nome not in acionistas:
        return jsonify({'success': False, 'error': 'Acionista não encontrado'})

    new_nome = data.get('nome', nome).strip()
    active = data.get('active', True)

    if new_nome != nome:
        # Renomear acionista
        del acionistas[nome]
        acionistas[new_nome] = {'active': active}
    else:
        acionistas[nome]['active'] = active

    save_acionistas(acionistas)
    return jsonify({'success': True, 'message': 'Acionista atualizado!'})

@app.route('/api/acionistas/<nome>', methods=['DELETE'])
@admin_required
def api_delete_acionista(nome):
    acionistas = load_acionistas()

    if nome not in acionistas:
        return jsonify({'success': False, 'error': 'Acionista não encontrado'})

    del acionistas[nome]
    save_acionistas(acionistas)
    return jsonify({'success': True, 'message': 'Acionista excluído!'})

# ==================================================================================
# ROTAS DE PROCESSAMENTO
# ==================================================================================

@app.route('/processar', methods=['POST'])
@login_required
def processar():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400

        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400

        if not arquivo.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Apenas arquivos .xlsx são permitidos'}), 400

        # Salva arquivo
        filename = secure_filename(arquivo.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        arquivo.save(filepath)

        # Processa
        resultado = processar_arquivo_rv_dsr(filepath)

        # Remove arquivo temporário
        os.remove(filepath)

        return jsonify({
            'success': True,
            **resultado
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download/<filename>')
@login_required
def download(filename):
    try:
        filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/download-zip', methods=['POST'])
@login_required
def download_zip():
    """Cria e baixa um ZIP com todos os arquivos gerados"""
    try:
        data = request.get_json()
        arquivos = data.get('arquivos', [])

        if not arquivos:
            return jsonify({'error': 'Nenhum arquivo para compactar'}), 400

        # Nome da pasta: data_rv_dsr (ex: 20260120_rv_dsr)
        data_hoje = datetime.now().strftime("%Y%m%d")
        nome_pasta = f"{data_hoje}_rv_dsr"
        nome_zip = f"{nome_pasta}.zip"

        # Caminho do ZIP
        zip_path = os.path.join(app.config['OUTPUT_FOLDER'], nome_zip)

        # Cria o ZIP
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for arquivo in arquivos:
                arquivo_path = os.path.join(app.config['OUTPUT_FOLDER'], arquivo)
                if os.path.exists(arquivo_path):
                    # Coloca o arquivo dentro de uma pasta no ZIP
                    zf.write(arquivo_path, os.path.join(nome_pasta, arquivo))

        return send_file(zip_path, as_attachment=True, download_name=nome_zip)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================================================================================
# ROTAS DE ROYALTIES
# ==================================================================================

@app.route('/api/royalties/config', methods=['GET'])
@login_required
def api_get_royalties_config():
    config = load_royalties_config()
    return jsonify(config)


@app.route('/api/royalties/produtos', methods=['POST'])
@login_required
def api_add_royalties_produto():
    data = request.get_json()
    produto = data.get('produto', '').strip().upper()

    if not produto:
        return jsonify({'success': False, 'error': 'Produto é obrigatório'})

    config = load_royalties_config()

    if produto in config['produtos_nao_royalties']:
        return jsonify({'success': False, 'error': 'Produto já existe na lista'})

    config['produtos_nao_royalties'].append(produto)
    save_royalties_config(config)
    return jsonify({'success': True, 'message': 'Produto adicionado!'})


@app.route('/api/royalties/produtos/<produto>', methods=['DELETE'])
@login_required
def api_delete_royalties_produto(produto):
    config = load_royalties_config()

    if produto not in config['produtos_nao_royalties']:
        return jsonify({'success': False, 'error': 'Produto não encontrado'})

    config['produtos_nao_royalties'].remove(produto)
    save_royalties_config(config)
    return jsonify({'success': True, 'message': 'Produto removido!'})


@app.route('/api/royalties/clientes', methods=['POST'])
@login_required
def api_add_royalties_cliente():
    data = request.get_json()
    cliente = data.get('cliente', '').strip().upper()

    if not cliente:
        return jsonify({'success': False, 'error': 'Cliente é obrigatório'})

    config = load_royalties_config()

    if cliente in config['clientes_nao_royalties']:
        return jsonify({'success': False, 'error': 'Cliente já existe na lista'})

    config['clientes_nao_royalties'].append(cliente)
    save_royalties_config(config)
    return jsonify({'success': True, 'message': 'Cliente adicionado!'})


@app.route('/api/royalties/clientes/<cliente>', methods=['DELETE'])
@login_required
def api_delete_royalties_cliente(cliente):
    config = load_royalties_config()

    if cliente not in config['clientes_nao_royalties']:
        return jsonify({'success': False, 'error': 'Cliente não encontrado'})

    config['clientes_nao_royalties'].remove(cliente)
    save_royalties_config(config)
    return jsonify({'success': True, 'message': 'Cliente removido!'})


def to_number_ptbr(value):
    """Converte valor brasileiro para float"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    s = s.replace('\u00a0', '').replace(' ', '')
    s = s.replace('.', '').replace(',', '.')

    try:
        return float(s)
    except:
        return 0.0


def to_number_ptbr_series(series):
    """Converte série pandas com valores brasileiros para numérico"""
    if pd.api.types.is_numeric_dtype(series):
        return series.fillna(0)
    s = series.astype(str).str.strip()
    s = s.str.replace('\u00a0', '', regex=False)
    s = s.str.replace(' ', '', regex=False)
    s = s.str.replace('.', '', regex=False)
    s = s.str.replace(',', '.', regex=False)
    return pd.to_numeric(s, errors='coerce').fillna(0)


def escolher_coluna_valor(df, palavras_chave):
    """Escolhe a melhor coluna de valor no DataFrame"""
    candidatos = []
    for col in df.columns:
        low = str(col).lower()
        if any(p in low for p in palavras_chave):
            candidatos.append(col)

    for col in candidatos:
        nums = to_number_ptbr_series(df[col])
        if nums.abs().sum() > 0:
            return col

    melhor = None
    melhor_score = -1.0
    for col in df.columns:
        nums = to_number_ptbr_series(df[col])
        score = float(nums.abs().sum())
        if score > melhor_score:
            melhor_score = score
            melhor = col

    return melhor if melhor_score > 0 else None


def escolher_campos_linhas(df):
    """Escolhe as melhores colunas para agrupar"""
    cols = df.columns.tolist()
    escolhidos = []

    for col in cols:
        low = str(col).lower()
        if ('cliente' in low) or ('cod' in low) or ('cód' in low) or ('codigo' in low):
            escolhidos.append(col)

    for col in cols:
        low = str(col).lower()
        if ('nome' in low) and ('cliente' not in low):
            escolhidos.append(col)

    seen = set()
    escolhidos = [c for c in escolhidos if not (c in seen or seen.add(c))]

    if len(escolhidos) >= 1:
        return escolhidos

    return cols[:2] if len(cols) >= 2 else cols[:1]


def criar_validacao_sim_nao(df, campos_linhas, campo_valor):
    """Cria tabela de validação agrupando por SIM/NÃO"""
    temp = df.copy()

    if 'Royalties' not in temp.columns:
        raise ValueError("Coluna 'Royalties' não existe no DataFrame.")
    if campo_valor is None or campo_valor not in temp.columns:
        raise ValueError("Não foi possível identificar a coluna de valor.")
    for c in campos_linhas:
        if c not in temp.columns:
            raise ValueError(f"Coluna '{c}' não existe no DataFrame.")

    temp['Royalties'] = temp['Royalties'].astype(str).str.strip().str.upper()
    temp['Royalties'] = temp['Royalties'].replace({'NAO': 'NÃO'})
    temp.loc[~temp['Royalties'].isin(['SIM', 'NÃO']), 'Royalties'] = 'SIM'

    temp[campo_valor] = to_number_ptbr_series(temp[campo_valor])

    agg = (
        temp.groupby(campos_linhas + ['Royalties'], dropna=False)[campo_valor]
        .sum()
        .unstack('Royalties', fill_value=0)
        .reset_index()
    )

    if 'SIM' not in agg.columns:
        agg['SIM'] = 0
    if 'NÃO' not in agg.columns:
        agg['NÃO'] = 0

    agg = agg[campos_linhas + ['NÃO', 'SIM']]
    agg['Total'] = agg['NÃO'] + agg['SIM']

    total_row = {c: 'Total Geral' for c in campos_linhas}
    total_row['NÃO'] = float(agg['NÃO'].sum())
    total_row['SIM'] = float(agg['SIM'].sum())
    total_row['Total'] = float(agg['Total'].sum())

    agg = pd.concat([agg, pd.DataFrame([total_row])], ignore_index=True)
    return agg


def processar_royalties(arquivo_path):
    """Processa arquivo Excel de royalties"""
    config = load_royalties_config()
    produtos_nao = set(p.strip().upper() for p in config['produtos_nao_royalties'])
    clientes_nao = set(c.strip().upper() for c in config['clientes_nao_royalties'])

    # Carrega arquivo
    wb_original = openpyxl.load_workbook(arquivo_path)
    xls = pd.ExcelFile(arquivo_path)
    todas_abas = {aba: pd.read_excel(arquivo_path, sheet_name=aba) for aba in xls.sheet_names}

    resultado = {
        'fat_sim': 0,
        'fat_nao': 0,
        'fat_total': 0,
        'baixas_sim': 0,
        'baixas_nao': 0,
        'baixas_total': 0
    }

    # Processa Detalhado NF
    if "Detalhado NF" in todas_abas:
        df_nf = todas_abas["Detalhado NF"].copy()

        # Adiciona coluna Royalties baseado na coluna H (índice 7)
        df_nf["Royalties"] = df_nf.iloc[:, 7].apply(
            lambda x: "NÃO" if str(x).strip().upper() in produtos_nao else "SIM"
        )
        todas_abas["Detalhado NF"] = df_nf

        # Calcula totais para validação
        valor_col = None
        for col in df_nf.columns:
            if any(p in str(col).lower() for p in ['total', 'valor']):
                valor_col = col
                break

        if valor_col:
            df_nf['_valor_num'] = df_nf[valor_col].apply(to_number_ptbr)
            resultado['fat_sim'] = float(df_nf[df_nf['Royalties'] == 'SIM']['_valor_num'].sum())
            resultado['fat_nao'] = float(df_nf[df_nf['Royalties'] == 'NÃO']['_valor_num'].sum())
            resultado['fat_total'] = resultado['fat_sim'] + resultado['fat_nao']

        # Gera VALIDAÇÃO FAT
        try:
            campos = escolher_campos_linhas(df_nf)
            valor = escolher_coluna_valor(df_nf, ['total', 'valor'])
            todas_abas['VALIDAÇÃO FAT'] = criar_validacao_sim_nao(df_nf, campos, valor)
        except Exception as e:
            todas_abas['VALIDAÇÃO FAT'] = pd.DataFrame([{'Erro': 'Falha ao criar VALIDAÇÃO FAT', 'Motivo': str(e)}])

    # Processa Detalhado Baixas
    if "Detalhado Baixas" in todas_abas:
        df_baixas = todas_abas["Detalhado Baixas"].copy()

        # Adiciona coluna Royalties baseado na coluna B (índice 1)
        df_baixas["Royalties"] = df_baixas.iloc[:, 1].apply(
            lambda x: "NÃO" if str(x).strip().upper() in clientes_nao else "SIM"
        )
        todas_abas["Detalhado Baixas"] = df_baixas

        # Calcula totais para validação
        valor_col = None
        for col in df_baixas.columns:
            if any(p in str(col).lower() for p in ['total', 'valor', 'baixa']):
                valor_col = col
                break

        if valor_col:
            df_baixas['_valor_num'] = df_baixas[valor_col].apply(to_number_ptbr)
            resultado['baixas_sim'] = float(df_baixas[df_baixas['Royalties'] == 'SIM']['_valor_num'].sum())
            resultado['baixas_nao'] = float(df_baixas[df_baixas['Royalties'] == 'NÃO']['_valor_num'].sum())
            resultado['baixas_total'] = resultado['baixas_sim'] + resultado['baixas_nao']

        # Gera VALIDAÇÃO BAIXAS
        try:
            campos = escolher_campos_linhas(df_baixas)
            valor = escolher_coluna_valor(df_baixas, ['total', 'valor', 'baixa'])
            todas_abas['VALIDAÇÃO BAIXAS'] = criar_validacao_sim_nao(df_baixas, campos, valor)
        except Exception as e:
            todas_abas['VALIDAÇÃO BAIXAS'] = pd.DataFrame([{'Erro': 'Falha ao criar VALIDAÇÃO BAIXAS', 'Motivo': str(e)}])

    # Gera arquivo de saída
    caminho = Path(arquivo_path)
    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_saida = f"{caminho.stem}_processado_{data_hora}{caminho.suffix}"
    arquivo_saida = os.path.join(ROYALTIES_OUTPUT_FOLDER, nome_saida)

    # Salva
    wb_original.save(arquivo_saida)
    wb = openpyxl.load_workbook(arquivo_saida)

    # Formatação
    header_fill = PatternFill(start_color="002233", end_color="002233", fill_type="solid")
    header_font = Font(bold=True, color="00DBFF", size=11)
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for nome_aba, df in todas_abas.items():
        # Remove coluna temporária se existir
        if '_valor_num' in df.columns:
            df = df.drop(columns=['_valor_num'])

        if nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(nome_aba)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, (datetime, pd.Timestamp)):
                    cell.number_format = "DD/MM/YYYY"

        # Aplica formatação
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_style
                cell.alignment = Alignment(vertical="center")
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0.00"

        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and cell.value != "":
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        ws.freeze_panes = ws["A2"]

    wb.save(arquivo_saida)

    resultado['arquivo_saida'] = nome_saida
    return resultado


@app.route('/api/royalties/processar', methods=['POST'])
@login_required
def api_processar_royalties():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400

        arquivo = request.files['arquivo']

        if arquivo.filename == '':
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400

        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Apenas arquivos .xlsx ou .xls são permitidos'}), 400

        # Salva arquivo temporário
        filename = secure_filename(arquivo.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        arquivo.save(filepath)

        # Processa
        resultado = processar_royalties(filepath)

        # Remove arquivo temporário
        os.remove(filepath)

        return jsonify({
            'success': True,
            **resultado
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': f"{str(e)}\n\n{traceback.format_exc()}"
        }), 500


@app.route('/api/royalties/download/<filename>')
@login_required
def api_download_royalties(filename):
    try:
        filepath = os.path.join(ROYALTIES_OUTPUT_FOLDER, filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404



# ==================================================================================
# ROTAS DE CONCILIACAO BANCARIA
# ==================================================================================

@app.route('/conciliacao')
@login_required
def conciliacao_menu():
    """Menu de opções de conciliação"""
    return render_template('conciliacao_menu.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })


@app.route('/conciliacao/bancaria-contabil')
@login_required
def conciliacao_bancaria_contabil():
    """Conciliação Bancária x Contábil"""
    return render_template('conciliacao_bancaria_contabil.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })


@app.route('/api/conciliacao/processar', methods=['POST'])
@login_required
def api_conciliacao_processar():
    """Processa conciliacao bancaria"""
    try:
        if processar_conciliacao is None:
            return jsonify({'success': False, 'error': 'Modulo de conciliacao nao encontrado'}), 500

        if 'arquivo_fin' not in request.files or 'arquivo_contabil' not in request.files:
            return jsonify({'success': False, 'error': 'Envie os dois arquivos (Financeiro e Contabil)'}), 400

        fin_file = request.files['arquivo_fin']
        cont_file = request.files['arquivo_contabil']

        if fin_file.filename == '' or cont_file.filename == '':
            return jsonify({'success': False, 'error': 'Selecione os 2 arquivos'}), 400

        # Salva arquivos temporarios
        fin_filename = secure_filename(fin_file.filename)
        cont_filename = secure_filename(cont_file.filename)

        fin_path = os.path.join(app.config['UPLOAD_FOLDER'], f"conc_fin_{fin_filename}")
        cont_path = os.path.join(app.config['UPLOAD_FOLDER'], f"conc_cont_{cont_filename}")

        fin_file.save(fin_path)
        cont_file.save(cont_path)

        # Gera arquivo de saida
        data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"conciliacao_{data_hora}.xlsx"
        output_path = os.path.join(CONCILIACAO_OUTPUT_FOLDER, output_filename)

        # Processa conciliacao
        stats = processar_conciliacao(
            financeiro_path=fin_path,
            contabil_path=cont_path,
            output_xlsx=output_path,
            tolerancia=0.01,
            min_len=3
        )

        # Remove arquivos temporarios
        try:
            os.remove(fin_path)
            os.remove(cont_path)
        except:
            pass

        return jsonify({
            'success': True,
            'arquivo': output_filename,
            **stats
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': f"{str(e)}\n\n{traceback.format_exc()}"
        }), 500


@app.route('/api/conciliacao/download/<filename>')
@login_required
def api_conciliacao_download(filename):
    """Download do arquivo de conciliacao"""
    try:
        filepath = os.path.join(CONCILIACAO_OUTPUT_FOLDER, filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


# ==================================================================================
# ROTAS DE CONCILIACAO CLIENTE (NFs x Recebimentos)
# ==================================================================================

@app.route('/conciliacao/cliente')
@login_required
def conciliacao_cliente():
    """Conciliação Cliente - NFs x Recebimentos"""
    return render_template('conciliacao_bancaria_cliente.html', user={
        'username': session.get('user'),
        'name': session.get('name'),
        'is_admin': session.get('is_admin')
    })


@app.route('/api/conciliacao-cliente/processar', methods=['POST'])
@login_required
def api_conciliacao_cliente_processar():
    """Processa conciliacao cliente"""
    try:
        if processar_conciliacao_cliente is None:
            return jsonify({'success': False, 'error': 'Modulo de conciliacao cliente nao encontrado'}), 500

        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'error': 'Envie o arquivo de Razão Contábil'}), 400

        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            return jsonify({'success': False, 'error': 'Selecione o arquivo'}), 400

        saldo_inicial = request.form.get('saldo_inicial', '0')
        data_inicio = request.form.get('data_inicio', '')
        data_fim = request.form.get('data_fim', '')

        # Salva arquivo temporario
        filename = secure_filename(arquivo.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"conc_cli_{filename}")
        arquivo.save(filepath)

        # Arquivo financeiro (opcional)
        filepath_fin = None
        if 'arquivo_financeiro' in request.files:
            arquivo_fin = request.files['arquivo_financeiro']
            if arquivo_fin.filename != '':
                fin_filename = secure_filename(arquivo_fin.filename)
                filepath_fin = os.path.join(app.config['UPLOAD_FOLDER'], f"conc_cli_fin_{fin_filename}")
                arquivo_fin.save(filepath_fin)

        # Gera arquivo de saida
        data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"conciliacao_cliente_{data_hora}.xlsx"
        output_path = os.path.join(CONCILIACAO_OUTPUT_FOLDER, output_filename)

        # Processa conciliacao
        resultado = processar_conciliacao_cliente(
            arquivo_path=filepath,
            saldo_inicial=saldo_inicial,
            data_inicio=data_inicio,
            data_fim=data_fim,
            output_path=output_path,
            arquivo_financeiro_path=filepath_fin
        )

        # Remove arquivos temporarios
        try:
            os.remove(filepath)
        except:
            pass
        if filepath_fin:
            try:
                os.remove(filepath_fin)
            except:
                pass

        return jsonify({
            'success': True,
            'arquivo': output_filename,
            **resultado
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': f"{str(e)}\n\n{traceback.format_exc()}"
        }), 500


@app.route('/api/conciliacao-cliente/download/<filename>')
@login_required
def api_conciliacao_cliente_download(filename):
    """Download do arquivo de conciliacao cliente"""
    try:
        filepath = os.path.join(CONCILIACAO_OUTPUT_FOLDER, filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


# ==================================================================================
# NOTAS DE SERVIÇOS TOMADOS — PROCESSADOR NFS-e PDF → TXT
# ==================================================================================

# Mapeamento município → (UF, IBGE) para os municípios mais comuns
_IBGE_MAP = {
    'FORTALEZA': ('CE', '2304400'),
    'SÃO PAULO': ('SP', '3550308'), 'SAOPAULO': ('SP', '3550308'),
    'RIO DE JANEIRO': ('RJ', '3304557'), 'RIODEJANEIRO': ('RJ', '3304557'),
    'BELO HORIZONTE': ('MG', '3106200'), 'BELOHORIZONTE': ('MG', '3106200'),
    'SALVADOR': ('BA', '2927408'),
    'CURITIBA': ('PR', '4106902'),
    'RECIFE': ('PE', '2611606'),
    'MANAUS': ('AM', '1302603'),
    'PORTO ALEGRE': ('RS', '4314902'), 'PORTOALEGRE': ('RS', '4314902'),
    'GOIÂNIA': ('GO', '5208707'), 'GOIANIA': ('GO', '5208707'),
    'BRASÍLIA': ('DF', '5300108'), 'BRASILIA': ('DF', '5300108'),
    'CAMPINAS': ('SP', '3509502'),
    'NATAL': ('RN', '2408102'),
    'MACEIÓ': ('AL', '2704302'), 'MACEIO': ('AL', '2704302'),
    'TERESINA': ('PI', '2211001'),
    'CAMPO GRANDE': ('MS', '5002704'), 'CAMPOGRANDE': ('MS', '5002704'),
    'CUIABÁ': ('MT', '5103403'), 'CUIABA': ('MT', '5103403'),
    'BELÉM': ('PA', '1501402'), 'BELEM': ('PA', '1501402'),
    'JOÃO PESSOA': ('PB', '2507507'), 'JOAOPESSOA': ('PB', '2507507'),
    'ARACAJU': ('SE', '2800308'),
    'PORTO VELHO': ('RO', '1100205'), 'PORTOVELHO': ('RO', '1100205'),
    'MACAPÁ': ('AP', '1600303'), 'MACAPA': ('AP', '1600303'),
    'BOA VISTA': ('RR', '1400100'), 'BOAVISTA': ('RR', '1400100'),
    'PALMAS': ('TO', '1721000'),
    'RIO BRANCO': ('AC', '1200401'), 'RIOBRANCO': ('AC', '1200401'),
    'FLORIANÓPOLIS': ('SC', '4205407'), 'FLORIANOPOLIS': ('SC', '4205407'),
    'VITÓRIA': ('ES', '3205309'), 'VITORIA': ('ES', '3205309'),
    # RJ
    'MARICÁ': ('RJ', '3302700'), 'MARICA': ('RJ', '3302700'),
    'NITERÓI': ('RJ', '3303302'), 'NITEROI': ('RJ', '3303302'),
    'DUQUE DE CAXIAS': ('RJ', '3301702'), 'DUQUEDECAXIAS': ('RJ', '3301702'),
    'NOVA IGUAÇU': ('RJ', '3303500'), 'NOVAIGUACU': ('RJ', '3303500'),
    'SÃO GONÇALO': ('RJ', '3304904'), 'SAOGONCALO': ('RJ', '3304904'),
    'CAMPOS DOS GOYTACAZES': ('RJ', '3301009'),
    'PETRÓPOLIS': ('RJ', '3303906'), 'PETROPOLIS': ('RJ', '3303906'),
    'VOLTA REDONDA': ('RJ', '3306305'), 'VOLTAREDONDA': ('RJ', '3306305'),
    'CABO FRIO': ('RJ', '3300704'), 'CABOFRIO': ('RJ', '3300704'),
    'ANGRA DOS REIS': ('RJ', '3300209'),
    'MACAÉ': ('RJ', '3302403'), 'MACAE': ('RJ', '3302403'),
    'NOVA FRIBURGO': ('RJ', '3303401'), 'NOVAFRIBURGO': ('RJ', '3303401'),
    'RESENDE': ('RJ', '3304201'),
    'TERESÓPOLIS': ('RJ', '3305206'), 'TERESOPOLIS': ('RJ', '3305206'),
    # SP adicionais
    'GUARULHOS': ('SP', '3518800'),
    'SÃO BERNARDO DO CAMPO': ('SP', '3548708'),
    'SANTO ANDRÉ': ('SP', '3547809'), 'SANTOANDRE': ('SP', '3547809'),
    'OSASCO': ('SP', '3534401'),
    'SOROCABA': ('SP', '3552205'),
    'RIBEIRÃO PRETO': ('SP', '3543402'), 'RIBEIRAOPRETO': ('SP', '3543402'),
    'SÃO JOSÉ DOS CAMPOS': ('SP', '3549904'),
    # MG adicionais
    'UBERLÂNDIA': ('MG', '3170206'), 'UBERLANDIA': ('MG', '3170206'),
    'CONTAGEM': ('MG', '3118601'),
    'JUIZ DE FORA': ('MG', '3136702'), 'JUIZDEFORA': ('MG', '3136702'),
}


def _so_digitos(txt):
    return re.sub(r'\D', '', txt or '')


def _valor_centavos(txt):
    """Converte 'R$1.234,56' ou '1.234,56' para centavos inteiros."""
    limpo = re.sub(r'[R$\s]', '', txt or '').replace('.', '').replace(',', '.')
    try:
        return int(round(float(limpo) * 100))
    except Exception:
        return 0


def _ibge_municipio(nome_municipio):
    """Retorna (uf, ibge) para um nome de município, ou ('','') se não encontrado."""
    if not nome_municipio:
        return '', ''

    def _norm(txt):
        txt = unicodedata.normalize('NFD', (txt or '').strip().upper())
        txt = ''.join(ch for ch in txt if unicodedata.category(ch) != 'Mn')
        txt = re.sub(r'[^A-Z0-9\s]', ' ', txt)
        txt = re.sub(r'\s+', ' ', txt).strip()
        return txt

    chave = _norm(nome_municipio)
    chave_sem_espaco = chave.replace(' ', '')

    for k, v in _IBGE_MAP.items():
        kn = _norm(k)
        if chave == kn or chave_sem_espaco == kn.replace(' ', ''):
            return v

    # Fallback para OCR truncado (ex.: "MARIC" -> "MARICA").
    if chave_sem_espaco:
        for k, v in _IBGE_MAP.items():
            kn = _norm(k).replace(' ', '')
            if len(chave_sem_espaco) >= 4 and (kn.startswith(chave_sem_espaco) or chave_sem_espaco.startswith(kn)):
                return v

    return '', ''


def _extrair_valor_por_rotulo(texto, rotulo_regex):
    """Busca valor monetário logo após o rótulo."""
    if not texto:
        return '0'

    rx_rotulo = re.compile(rotulo_regex, re.IGNORECASE)
    rx_valor = re.compile(r'R?\$?\s*([\d.]+,\d{2}|[\d]+\.\d{2})')

    texto_norm = re.sub(r'\s+', ' ', texto)
    for m_rot in rx_rotulo.finditer(texto_norm):
        trecho = texto_norm[m_rot.end():m_rot.end() + 120]
        m_val = rx_valor.search(trecho)
        if m_val:
            return str(_valor_centavos(m_val.group(1)))
    return '0'


def _extrair_tributo_por_linha(texto, rotulos):
    """Extrai valor de tributo buscando rótulo + valor monetário na mesma linha/janela curta."""
    if not texto:
        return '0'

    if isinstance(rotulos, str):
        rotulos = [rotulos]
    rx_label = re.compile('|'.join(rotulos), re.IGNORECASE)
    rx_valor = re.compile(r'R?\$?\s*([\d.]+,\d{2}|[\d]+\.\d{2})')

    linhas = [re.sub(r'\s+', ' ', ln).strip() for ln in texto.splitlines() if ln.strip()]
    for i, ln in enumerate(linhas):
        m_label = rx_label.search(ln)
        if not m_label:
            continue
        # Mesmo que haja vários tributos na mesma linha, pega o valor após o rótulo atual.
        m = rx_valor.search(ln[m_label.end():])
        if m:
            return str(_valor_centavos(m.group(1)))
        if i + 1 < len(linhas):
            janela = f"{ln} {linhas[i+1]}"
            m = rx_valor.search(janela[m_label.end():])
            if m:
                return str(_valor_centavos(m.group(1)))
    return '0'


def _extrair_numero_nota(texto):
    """Extrai número da NFS-e com heurísticas para evitar pegar códigos indevidos."""
    if not texto:
        return ''

    def _candidatos(snippet):
        brutos = []
        brutos.extend(re.findall(r'\b(\d{5,9})\b', snippet or ''))
        brutos.extend(re.findall(r'\b(\d{3}[ .-]\d{3,6})\b', snippet or ''))
        nums = []
        for b in brutos:
            n = re.sub(r'\D', '', b)
            if 5 <= len(n) <= 9:
                nums.append(n)
        # Evita anos e números improváveis
        return [n for n in nums if n not in ('2020', '2021', '2022', '2023', '2024', '2025', '2026', '2027')]

    cand = []

    # Tentativa mais forte: trecho entre "Número da NFS-e" e "Série da DPS"
    m_bloco = re.search(r'N[uú]mero\s*da\s*NFS-e([\s\S]{0,180}?)S[eé]rie\s*da\s*DPS', texto, re.IGNORECASE)
    if m_bloco:
        bloco_cands = _candidatos(m_bloco.group(1))
        if bloco_cands:
            # Regra principal: primeiro candidato no bloco específico do cabeçalho.
            return bloco_cands[0]

    # Tentativas diretas próximas ao rótulo
    for m in re.finditer(r'N[uú]mero\s*da\s*NFS-e', texto, re.IGNORECASE):
        snip = texto[m.end():m.end() + 120]
        for n in _candidatos(snip):
            cand.append((n, m.start()))

    # Padrão "NFS-e nº 123456"
    for m in re.finditer(r'NFS-e\s*(?:n[º°o]\s*)?(\d{5,9})\b', texto, re.IGNORECASE):
        cand.append((m.group(1), m.start()))

    # Linhas com NFS-e (evita linha de DPS)
    for ln in texto.splitlines():
        if not re.search(r'NFS-e', ln, re.IGNORECASE):
            continue
        if re.search(r'DPS', ln, re.IGNORECASE):
            continue
        for n in _candidatos(ln):
            cand.append((n, texto.find(ln)))

    if not cand:
        return ''

    # Fallback: prioriza maior tamanho e ocorrência mais cedo (mais perto do cabeçalho).
    cand.sort(key=lambda x: (-len(x[0]), x[1]))
    return cand[0][0]


def _parse_endereco_por_virgula(endereco):
    """
    Regra:
    - 1º trecho: rua
    - 2º trecho: numero
    - se houver 4+ trechos: 3º = complemento, 4º = bairro
    - se houver 3 trechos: 3º = bairro
    """
    partes = [p.strip() for p in (endereco or '').split(',')]
    partes = [p for p in partes if p != '']

    rua = partes[0] if len(partes) >= 1 else ''
    numero = partes[1] if len(partes) >= 2 else ''
    complemento = ''
    bairro = ''

    if len(partes) >= 4:
        complemento = partes[2]
        bairro = partes[3]
    elif len(partes) == 3:
        bairro = partes[2]

    return rua, numero, complemento, bairro


def _extrair_tributo_segmentado(texto, rotulo, proximos_rotulos):
    """Extrai tributo no trecho entre o rótulo atual e o próximo rótulo conhecido."""
    if not texto:
        return '0'
    pad_inicio = re.compile(rotulo, re.IGNORECASE)
    pad_fim = re.compile('|'.join(proximos_rotulos), re.IGNORECASE) if proximos_rotulos else None
    pad_valor = re.compile(r'R?\$?\s*([\d.]+,\d{2}|[\d]+\.\d{2})')

    t = re.sub(r'\s+', ' ', texto)
    m_ini = pad_inicio.search(t)
    if not m_ini:
        return '0'
    trecho = t[m_ini.end():]
    if pad_fim:
        m_fim = pad_fim.search(trecho)
        if m_fim:
            trecho = trecho[:m_fim.start()]
    m_val = pad_valor.search(trecho)
    return str(_valor_centavos(m_val.group(1))) if m_val else '0'


def _chars_para_texto(chars):
    """Converte lista de caracteres posicionados em texto, inserindo espaços pelo gap."""
    if not chars:
        return ''
    chars_ord = sorted(chars, key=lambda c: c['x0'])
    texto = chars_ord[0]['text']
    for i in range(1, len(chars_ord)):
        prev = chars_ord[i - 1]
        curr = chars_ord[i]
        x1_prev = prev.get('x1', prev['x0'] + abs(prev.get('width', 5)))
        gap = curr['x0'] - x1_prev
        ref_w = (abs(prev.get('width', 5)) + abs(curr.get('width', 5))) / 2
        if ref_w <= 0:
            ref_w = 5
        if gap > ref_w * 0.3:
            texto += ' '
        texto += curr['text']
    return texto


def _extrair_texto_nfs(pdf_path):
    """Extrai texto de NFS-e preservando espaços via posicionamento de caracteres."""
    paginas = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            chars = [c for c in page.chars if c.get('text', '').strip()]
            if not chars:
                paginas.append('')
                continue
            linhas = []
            linha_atual = []
            y_atual = None
            for char in sorted(chars, key=lambda c: (c['top'], c['x0'])):
                y = char['top']
                if y_atual is None or abs(y - y_atual) > 3:
                    if linha_atual:
                        linhas.append(_chars_para_texto(linha_atual))
                    linha_atual = [char]
                    y_atual = y
                else:
                    linha_atual.append(char)
            if linha_atual:
                linhas.append(_chars_para_texto(linha_atual))
            paginas.append('\n'.join(l for l in linhas if l.strip()))
    return '\n'.join(paginas)


def processar_nfs_pdf(pdf_path):
    """Extrai campos de uma NFS-e (PDF) e retorna dict com chaves coluna_*.

    Usa pdfplumber.extract_text() (mais simples e confiável que extração
    char-a-char) e retorna valores tipados: float para monetários/alíquota,
    int para ISSQN/série/mês/ano, str para os demais.
    montar_linha_nfs_txt() converte para o formato TXT do TOTVS.
    """
    with pdfplumber.open(pdf_path) as pdf:
        texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
    t = texto

    def _lim(v):
        return re.sub(r'\s+', ' ', v).strip() if v else ""

    def _to_float(v):
        if not v or str(v).strip() in ("-", ""):
            return None
        s = re.sub(r'[R$\s]', '', str(v)).replace('.', '').replace(',', '.')
        try:
            return float(s)
        except ValueError:
            return None

    # ── Número da nota ──────────────────────────────────────────────────
    num_nota = _extrair_numero_nota(t)

    # ── Série ───────────────────────────────────────────────────────────
    m = re.search(r'S[eé]rie\s*da\s*DPS\s+(\d+)', t, re.IGNORECASE)
    if not m:
        m = re.search(r'N[uú]mero\s*da\s*DPS\s+S[eé]rie\s*da\s*DPS[\s\S]{0,80}?\n\d+\s+(\d+)', t, re.IGNORECASE)
    serie = m.group(1) if m else '1'

    # ── Data emissão e competência ──────────────────────────────────────
    datas = re.findall(r'\b(\d{2}/\d{2}/\d{4})\b', t)
    data_emissao = datas[0] if datas else ""
    mes_comp = int(data_emissao[3:5]) if data_emissao else None
    ano_comp = int(data_emissao[6:10]) if data_emissao else None
    m_comp = re.search(r'Compet[eê]ncia\s*da\s*NFS-e\s+(\d{2}/\d{2}/\d{4})', t, re.IGNORECASE)
    if m_comp:
        comp = m_comp.group(1)
        mes_comp = int(comp[3:5])
        ano_comp = int(comp[6:10])

    # ── Seção emitente ──────────────────────────────────────────────────
    m_emit = re.search(r'EMITENTE\s*DA\s*NFS-e(.*?)TOMADOR\s*DO\s*SERVI', t, re.DOTALL | re.IGNORECASE)
    sec_emit = m_emit.group(1) if m_emit else t

    # ── CNPJ ────────────────────────────────────────────────────────────
    m = re.search(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})', sec_emit)
    cnpj = _so_digitos(m.group(1)) if m else ""

    # ── Razão social ────────────────────────────────────────────────────
    razao = ""
    m_nome = re.search(r'Nome/Nome\s*Empresarial(?:\s+E-mail)?\s*\n([\s\S]+?)(?=Endere[çc]o)', sec_emit, re.IGNORECASE)
    if m_nome:
        for linha in m_nome.group(1).splitlines():
            linha = linha.strip()
            if not linha or '@' in linha:
                continue
            if re.fullmatch(r'(E-mail|Endere[çc]o|Munic[íi]pio|CEP|Inscri[çc][ãa]o\s*Municipal)', linha, re.I):
                continue
            if len(linha) > 5:
                razao = linha
                break
    if not razao:
        m = re.search(r'([A-ZÁÉÍÓÚÃÕÊ][A-ZÁÉÍÓÚÃÕÊ\s&.,/]{4,}(?:LTDA|S\.?A\.?|ME|EPP|EIRELI|S/A)\.?)', sec_emit)
        razao = m.group(1).strip() if m else ""

    # ── Telefone e e-mail ────────────────────────────────────────────────
    m = re.search(r'\((\d{2})\)\s*(\d{4,5})-?(\d{4})', sec_emit)
    if m:
        telefone = m.group(1) + m.group(2) + m.group(3)
    else:
        m = re.search(r'\b(\d{10,11})\b', sec_emit)
        telefone = m.group(1) if m else ""
    m = re.search(r'[\w.+\-]+@[\w.\-]+\.[a-zA-Z]{2,}', sec_emit)
    email = m.group(0).lower() if m else ""

    # ── Endereço, UF, IBGE, CEP emitente ────────────────────────────────
    rua = numero = complemento = bairro = ""
    uf_emit = ibge_emit = cep_emit = ""
    m_cep = re.search(r'\b(\d{5}-\d{3})\b', sec_emit)
    if m_cep:
        cep_emit = _so_digitos(m_cep.group(1))
        # Prioriza linha com o rótulo Endereço logo acima do CEP
        m_linha_end = re.search(r'Endere[çc]o[^\n\r]*\n([^\n\r]+)', sec_emit, re.IGNORECASE)
        if m_linha_end:
            linha_end = m_linha_end.group(1).strip()
            if cep_emit in linha_end or m_cep.group(1) in linha_end or ',' in linha_end:
                ll = re.sub(r'\s*\d{5}-?\d{3}\s*.*$', '', linha_end).strip()
                ll = re.sub(r'^(Endere[çc]o|Logradouro)\s*[:\-]?\s*', '', ll, flags=re.IGNORECASE).strip()
                rua, numero, complemento, bairro = _parse_endereco_por_virgula(ll)
        # Busca linha que contenha o CEP para extrair cidade/UF
        for linha in sec_emit.splitlines():
            if m_cep.group(1) not in linha:
                continue
            ll = re.sub(r'\s*\d{5}-?\d{3}\s*.*$', '', linha).strip()
            ll = re.sub(r'\s+coluna\s+\w+', '', ll, flags=re.IGNORECASE).strip()
            ll = re.sub(r'^(Endere[çc]o|Logradouro)\s*[:\-]?\s*', '', ll, flags=re.IGNORECASE).strip()
            m_cidade = re.search(r',?\s*([A-Za-zÀ-ú]+(?:\s+[A-Za-zÀ-ú]+)*)\s*[-–/]\s*([A-Z]{2})\s*$', ll)
            if m_cidade:
                uf_emit = m_cidade.group(2)
                ibge_emit = _ibge_municipio(m_cidade.group(1).strip())[1]
                ll = ll[:m_cidade.start()].strip()
            if not rua:
                rua, numero, complemento, bairro = _parse_endereco_por_virgula(ll)
            break
    if not uf_emit:
        m = re.search(r'\b([A-Za-zÀ-ú]+(?:\s+[A-Za-zÀ-ú]+)*)\s*[-–/]\s*([A-Z]{2})\b', sec_emit)
        if m:
            uf_emit = m.group(2)
            ibge_emit = _ibge_municipio(m.group(1))[1]

    # ── Alíquota (guardada como float %, ex.: 5.0) ───────────────────────
    tn = unicodedata.normalize('NFD', t)
    tn = ''.join(c for c in tn if unicodedata.category(c) != 'Mn')
    tn = re.sub(r'\s+', ' ', tn)
    aliquota = 2.0
    m = re.search(r'aliquota\s*aplicada[^\d]{0,30}(\d{1,2}(?:[.,]\d{1,4})?)\s*%?', tn, re.IGNORECASE)
    if not m:
        m = re.search(r'aliquota[^\d]{0,30}(\d{1,2}(?:[.,]\d{1,4})?)\s*%?', tn, re.IGNORECASE)
    if not m:
        m = re.search(r'\b(\d{1,2}(?:[.,]\d{1,4})?)\s*%\s*(?:de\s*)?iss', tn, re.IGNORECASE)
    if not m:
        m = re.search(r'\biss(?:qn)?[^\d]{0,30}(\d{1,2}(?:[.,]\d{1,4})?)\s*%', tn, re.IGNORECASE)
    if m:
        try:
            aliquota = float(m.group(1).replace(',', '.'))
        except Exception:
            pass

    # ── Descrição do serviço ─────────────────────────────────────────────
    m = re.search(r'Descri[çc][ãa]o\s*do\s*Servi[çc]o\s*([\s\S]*?)(?=TRIBUTA[ÇC])', t, re.IGNORECASE)
    if m:
        descr = re.sub(r'\s*coluna\s+\w+', '', m.group(1), flags=re.IGNORECASE)
        descr = re.split(r'Reten[çc][ãa]o|Reten[çc][õo]es', descr, flags=re.IGNORECASE)[0]
        descr = re.sub(r'[+\-]', '', descr)
        descricao = ' '.join(descr.split()).strip()
    else:
        descricao = ""

    # ── Local de prestação ───────────────────────────────────────────────
    uf_local = ibge_local = ""
    m = re.search(r'Local\s*da\s*Presta[çc][ãa]o[\s\S]{0,120}?([A-Za-zÀ-ú\s]+?)\s*[-–/]\s*([A-Z]{2})\b', t, re.IGNORECASE)
    if m:
        uf_local = m.group(2)
        ibge_local = _ibge_municipio(m.group(1).strip())[1]
    if not uf_local:
        uf_local = uf_emit
        ibge_local = ibge_emit

    # ── Valor do serviço (float R$) ──────────────────────────────────────
    valor = None
    m = re.search(r'Valor\s*do\s*Servi[çc]o\s+Desconto\s*Condicionado[\s\S]{0,60}?\n\s*R\$\s*([\d.,]+)', t, re.IGNORECASE)
    if not m:
        m = re.search(r'Valor\s*do\s*Servi[çc]o\s+R\$\s*([\d.,]+)', t, re.IGNORECASE)
    if not m:
        m = re.search(r'Valor\s*do\s*Servi[çc]o\D{0,10}([\d.,]+)', t, re.IGNORECASE)
    if m:
        valor = _to_float(m.group(1))

    # ── Tributos federais (float R$) ─────────────────────────────────────
    m_fed = re.search(r'TRIBUTA[ÇC][ÃA]O\s*FEDERAL([\s\S]*?)(?:VALOR\s*TOTAL\s*DA\s*NFS|TOTAIS\s*APROXIMADOS|$)', t, re.IGNORECASE)
    sec_fed = m_fed.group(1) if m_fed else t

    def _trib(rotulo, proximos):
        v = _extrair_tributo_segmentado(sec_fed, rotulo, proximos)
        if not v or v == '0':
            return None
        try:
            return int(v) / 100.0
        except Exception:
            return None

    irrf   = _trib(r'\bIRRF\b',   [r'\bPIS\b', r'\bC[O0]FINS\b', r'Contribui[çc][oõ]es\s*Sociais', r'\bCSLL\b'])
    pis    = _trib(r'\bPIS\b',    [r'\bC[O0]FINS\b', r'Contribui[çc][oõ]es\s*Sociais', r'\bCSLL\b', r'\bIRRF\b'])
    cofins = _trib(r'\bC[O0]FINS\b', [r'Contribui[çc][oõ]es\s*Sociais', r'\bCSLL\b', r'\bIRRF\b', r'\bPIS\b'])
    csll   = _trib(r'Contribui[çc][oõ]es\s*Sociais|\bCSLL\b', [r'\bIRRF\b', r'\bPIS\b', r'\bC[O0]FINS\b'])

    # ── ISSQN retido ────────────────────────────────────────────────────
    m = re.search(r'Reten[çc][ãa]o\s*do\s*ISSQN\s+(N[ãa]o\s*Retido|Retido)', t, re.IGNORECASE)
    issqn_retido = 0 if (m and re.search(r'N[ãa]o', m.group(1), re.I)) or not m else 1

    return {
        # Emitente
        'coluna_D_CNPJ':                cnpj,
        'coluna_E_razao_social':        _lim(razao),
        'coluna_H_UF':                  uf_emit,
        'coluna_I_cod_ibge':            ibge_emit,
        'coluna_J_CEP':                 cep_emit,
        'coluna_K_rua':                 rua,
        'coluna_L_numero':              numero,
        'coluna_M_complemento':         complemento,
        'coluna_N_bairro':              bairro,
        'coluna_O_telefone':            telefone,
        'coluna_P_email':               email,
        # Nota
        'coluna_R_numero_nota':         num_nota,
        'coluna_S_serie':               serie,
        'coluna_T_data_emissao':        data_emissao,
        'coluna_V_mes_competencia':     mes_comp,
        'coluna_W_ano_competencia':     ano_comp,
        # Serviço
        'coluna_X_cod_tributacao':      '620400001',
        'coluna_Y_aliquota_pct':        aliquota,
        'coluna_Z_descricao_servico':   descricao,
        'coluna_AB_UF_prestacao':       uf_local,
        'coluna_AC_cod_ibge_prestacao': ibge_local,
        # Valores (float R$; montar_linha_nfs_txt converte para centavos)
        'coluna_AG_valor_nota':         valor,
        'coluna_AL_IRRF':               irrf,
        'coluna_AM_PIS':                pis,
        'coluna_AN_COFINS':             cofins,
        'coluna_AO_csll':               csll,
        'coluna_AR_ISSQN_retido':       issqn_retido,
    }


def _diagnostico_nfs_campos(dados):
    """Gera diagnóstico simples de campos-chave para facilitar depuração."""
    campos_chave = [
        'coluna_D_CNPJ', 'coluna_E_razao_social',
        'coluna_H_UF', 'coluna_I_cod_ibge', 'coluna_J_CEP',
        'coluna_K_rua', 'coluna_L_numero', 'coluna_N_bairro',
        'coluna_T_data_emissao', 'coluna_X_cod_tributacao',
        'coluna_Y_aliquota_pct', 'coluna_Z_descricao_servico',
        'coluna_AB_UF_prestacao', 'coluna_AC_cod_ibge_prestacao',
        'coluna_AG_valor_nota', 'coluna_R_numero_nota',
    ]
    faltando = [k for k in campos_chave if not str(dados.get(k) or '').strip()]
    return {
        'faltando': faltando,
        'ok': len(faltando) == 0,
    }


def raspar_nfs_para_json(pdf_path, arquivo_origem=''):
    """Extrai dados da NFS e retorna estrutura JSON de apoio ao TXT."""
    dados = processar_nfs_pdf(pdf_path)
    return {
        'arquivo': arquivo_origem or os.path.basename(pdf_path),
        'extraido_em': datetime.now().isoformat(timespec='seconds'),
        'campos': dados,
        'diagnostico': _diagnostico_nfs_campos(dados),
    }

def montar_linha_nfs_txt(d):
    """Monta a linha no formato TXT para importação no TOTVS.

    Aceita dict com chaves coluna_* (saída de processar_nfs_pdf).
    Valores monetários e alíquota são float R$/% e são convertidos
    para centavos/centésimos inteiros aqui.
    """
    def _s(key, default=''):
        v = d.get(key)
        return str(v) if v is not None else default

    def _cents(key):
        """Multiplica por 100 e arredonda; retorna '0' se ausente/None."""
        v = d.get(key)
        if v is None:
            return '0'
        try:
            return str(int(round(float(v) * 100)))
        except Exception:
            return '0'

    campos = [
        '2.0',                                         # 1  versão
        '2',                                           # 2  fixo
        '2',                                           # 3  fixo
        _s('coluna_D_CNPJ'),                           # 4  CNPJ (só dígitos)
        _s('coluna_E_razao_social'),                   # 5  Razão Social
        '0',                                           # 6  fixo
        '1058',                                        # 7  fixo (cód. interno)
        _s('coluna_H_UF'),                             # 8  UF do prestador
        _s('coluna_I_cod_ibge'),                       # 9  IBGE do prestador
        _s('coluna_J_CEP'),                            # 10 CEP (só dígitos)
        _s('coluna_K_rua'),                            # 11 Logradouro
        _s('coluna_L_numero'),                         # 12 Número
        _s('coluna_M_complemento'),                    # 13 Complemento
        _s('coluna_N_bairro'),                         # 14 Bairro
        _s('coluna_O_telefone'),                       # 15 Telefone (só dígitos)
        _s('coluna_P_email'),                          # 16 E-mail
        '7',                                           # 17 fixo
        '67',                                          # 18 fixo
        '',                                            # 19 vazio
        _s('coluna_T_data_emissao'),                   # 20 Data emissão DD/MM/AAAA
        _s('coluna_S_serie', '1'),                     # 21 Série
        _s('coluna_V_mes_competencia'),                # 22 Mês (MM)
        _s('coluna_W_ano_competencia'),                # 23 Ano (AAAA)
        _s('coluna_X_cod_tributacao', '620400001'),    # 24 Cód. tributação nacional
        _cents('coluna_Y_aliquota_pct'),               # 25 Alíquota × 100 (centésimos)
        _s('coluna_Z_descricao_servico'),              # 26 Descrição do serviço
        '1058',                                        # 27 fixo (local prestação)
        _s('coluna_AB_UF_prestacao'),                  # 28 UF local de prestação
        _s('coluna_AC_cod_ibge_prestacao'),            # 29 IBGE local de prestação
        '2',                                           # 30 fixo
        '',                                            # 31 vazio
        '',                                            # 32 vazio
        _cents('coluna_AG_valor_nota'),                # 33 Valor serviço (centavos)
        '', '', '', '', '',                             # 34-38 vazios
        _cents('coluna_AL_IRRF'),                      # 39 IRRF (centavos)
        _cents('coluna_AM_PIS'),                       # 40 PIS (centavos)
        _cents('coluna_AN_COFINS'),                    # 41 COFINS (centavos)
        _cents('coluna_AO_csll'),                      # 42 CSLL (centavos)
        '', '',                                        # 43-44 vazios
        _s('coluna_AR_ISSQN_retido', '0'),             # 45 ISSQN retido (0=não)
        '',                                            # 46 vazio
        _s('coluna_R_numero_nota'),                    # 47 Número da NFS-e
    ]
    return ';'.join(campos)


@app.route('/notas-servicos')
@login_required
def notas_servicos():
    return render_template('notas_servicos.html', user=session.get('user'))


@app.route('/api/notas-servicos/processar', methods=['POST'])
@login_required
def api_notas_servicos_processar():
    arquivos = request.files.getlist('arquivos')
    if not arquivos or not any(f.filename for f in arquivos):
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    resultados = []
    resultados_json = []
    linhas_txt = []
    erros = []

    for arq in arquivos:
        if not arq.filename.lower().endswith('.pdf'):
            erros.append(f'{arq.filename}: não é um PDF')
            continue

        filename = secure_filename(arq.filename)
        tmp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        arq.save(tmp_path)

        try:
            nota_json = raspar_nfs_para_json(tmp_path, arquivo_origem=arq.filename)
            dados = nota_json['campos']
            dados['arquivo_origem'] = arq.filename
            linha = montar_linha_nfs_txt(dados)
            linhas_txt.append(linha)
            resultados.append({'arquivo': arq.filename, 'dados': dados, 'linha': linha})
            resultados_json.append(nota_json)
        except Exception as e:
            erros.append(f'{arq.filename}: {str(e)}')
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass

    if not linhas_txt:
        return jsonify({'error': 'Nenhuma nota processada. ' + '; '.join(erros)}), 400

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nome_txt = f'notas_servicos_{timestamp}.txt'
    path_txt = os.path.join(NOTAS_OUTPUT_FOLDER, nome_txt)
    with open(path_txt, 'w', encoding='utf-8') as f:
        f.write('\n'.join(linhas_txt))

    nome_json = f'notas_servicos_{timestamp}.json'
    path_json = os.path.join(NOTAS_OUTPUT_FOLDER, nome_json)
    with open(path_json, 'w', encoding='utf-8') as f:
        json.dump({
            'gerado_em': datetime.now().isoformat(timespec='seconds'),
            'total_notas': len(resultados_json),
            'notas': resultados_json,
        }, f, ensure_ascii=False, indent=2)

    return jsonify({
        'success': True,
        'total': len(linhas_txt),
        'erros': erros,
        'arquivo': nome_txt,
        'arquivo_json': nome_json,
        'resultados': resultados,
    })


@app.route('/api/notas-servicos/download/<filename>')
@login_required
def api_notas_servicos_download(filename):
    try:
        filepath = os.path.join(NOTAS_OUTPUT_FOLDER, filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


if __name__ == '__main__':
    app.run(debug=True, port=5001, use_reloader=False)
