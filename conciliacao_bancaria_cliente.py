"""
Sistema de Conciliação Contábil - TOTVS
Desenvolvido por: Albe
Versão: 1.0

Processa planilhas de razão contábil e realiza conciliação automática
de NFs e Recebimentos com geração de relatórios detalhados.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
from decimal import Decimal
from collections import defaultdict
import os
import sys


class ConciliacaoCliente:
    def __init__(self):
        self.workbook_input = None
        self.workbook_output = None
        self.saldo_inicial = Decimal('0')
        self.data_inicio = None
        self.data_fim = None
        self.lancamentos = []
        self.nfs = {}  # {numero_nf: {dados}}
        self.recebimentos = {}  # {numero_rec: {dados}}
        self.matches = []
        self.nao_encontrados_nf = []
        self.nao_encontrados_rec = []
        
    def iniciar(self):
        """Método principal que executa todo o processo"""
        self.exibir_banner()
        
        try:
            # 1. Carregar planilha
            caminho_arquivo = self.solicitar_arquivo()
            self.carregar_planilha(caminho_arquivo)
            
            # 2. Coletar parâmetros
            self.coletar_parametros()
            
            # 3. Processar lançamentos
            print("\n⏳ Processando lançamentos...")
            self.processar_lancamentos()
            
            # 4. Realizar matching
            print("⏳ Realizando matching...")
            self.realizar_matching()
            
            # 5. Gerar relatório
            print("⏳ Gerando relatório...")
            nome_saida = self.gerar_relatorio(caminho_arquivo)
            
            # 6. Exibir resumo
            self.exibir_resumo()
            
            print(f"\n✓ Processamento concluído com sucesso!")
            print(f"📄 Arquivo gerado: {nome_saida}\n")
            
        except Exception as e:
            print(f"\n✗ Erro: {str(e)}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def exibir_banner(self):
        """Exibe banner do sistema"""
        print("\n" + "="*60)
        print("    Sistema de Conciliação Contábil - TOTVS")
        print("    Desenvolvido por: Albe")
        print("="*60 + "\n")
    
    def solicitar_arquivo(self):
        """Solicita o caminho do arquivo de entrada"""
        caminho = input("📁 Caminho da planilha de entrada: ").strip()
        
        if not os.path.exists(caminho):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
        
        return caminho
    
    def carregar_planilha(self, caminho):
        """Carrega a planilha Excel"""
        print("\n⏳ Carregando planilha...")
        self.workbook_input = openpyxl.load_workbook(caminho, data_only=True)
        print("✓ Planilha carregada")
    
    def coletar_parametros(self):
        """Coleta parâmetros do usuário"""
        print("\n📋 Parâmetros de Conciliação:\n")
        
        # Saldo Inicial
        while True:
            try:
                saldo_str = input("💰 Saldo Inicial (R$): ").strip()
                self.saldo_inicial = self.parse_valor(saldo_str)
                break
            except:
                print("⚠️  Valor inválido. Use formato: 1000.50 ou 1.000,50")
        
        # Data Início
        while True:
            try:
                data_inicio_str = input("📅 Data Início (DD/MM/YYYY): ").strip()
                self.data_inicio = datetime.strptime(data_inicio_str, "%d/%m/%Y")
                break
            except:
                print("⚠️  Data inválida. Use formato: DD/MM/YYYY")
        
        # Data Fim
        while True:
            try:
                data_fim_str = input("📅 Data Fim (DD/MM/YYYY): ").strip()
                self.data_fim = datetime.strptime(data_fim_str, "%d/%m/%Y")
                break
            except:
                print("⚠️  Data inválida. Use formato: DD/MM/YYYY")
        
        print(f"\n✓ Intervalo: {self.data_inicio.strftime('%d/%m/%Y')} até {self.data_fim.strftime('%d/%m/%Y')}")
    
    def parse_valor(self, valor):
        """Converte string de valor para Decimal"""
        if valor is None or valor == '':
            return Decimal('0')
        
        # Se já for número
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        
        # Converter string
        valor_str = str(valor).strip()
        
        # Remover símbolo de moeda e espaços
        valor_str = valor_str.replace('R$', '').replace(' ', '')
        
        # Verificar se termina com D ou C
        multiplicador = Decimal('1')
        if valor_str.endswith('D'):
            valor_str = valor_str[:-1].strip()
            multiplicador = Decimal('1')
        elif valor_str.endswith('C'):
            valor_str = valor_str[:-1].strip()
            multiplicador = Decimal('-1')
        
        # Substituir separadores
        # Se tiver ponto e vírgula, assumir formato brasileiro (1.000,50)
        if ',' in valor_str and '.' in valor_str:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        # Se tiver apenas vírgula, assumir formato brasileiro (1000,50)
        elif ',' in valor_str:
            valor_str = valor_str.replace(',', '.')
        
        try:
            return Decimal(valor_str) * multiplicador
        except:
            return Decimal('0')
    
    def esta_no_intervalo(self, data):
        """Verifica se a data está no intervalo especificado"""
        if data is None:
            return False
        
        # Se já for datetime
        if isinstance(data, datetime):
            data_comparacao = data
        else:
            # Tentar converter
            try:
                if isinstance(data, str):
                    data_comparacao = datetime.strptime(data, "%Y-%m-%d %H:%M:%S")
                else:
                    return False
            except:
                return False
        
        return self.data_inicio <= data_comparacao <= self.data_fim
    
    def processar_lancamentos(self):
        """Processa todos os lançamentos da planilha"""
        # Buscar aba de lançamentos
        ws = None
        for sheet_name in self.workbook_input.sheetnames:
            if 'Lançamentos Contábeis' in sheet_name or 'Lancamentos Contabeis' in sheet_name:
                ws = self.workbook_input[sheet_name]
                break
        
        if ws is None:
            abas = ', '.join(self.workbook_input.sheetnames)
            raise Exception(
                f"Aba de Lançamentos Contábeis não encontrada!\n\n"
                f"O sistema procura uma aba com 'Lançamentos Contábeis' no nome.\n"
                f"Abas encontradas no arquivo: {abas}\n\n"
                f"Verifique se você selecionou o arquivo correto no campo 'Razão Contábil'."
            )
        
        # Cabeçalho está na linha 2, dados começam na linha 3
        header_row = 2
        data_start_row = 3
        
        total_processados = 0
        
        for row_num in range(data_start_row, ws.max_row + 1):
            # Ler células
            data = ws.cell(row_num, 2).value  # Coluna B - DATA
            lote = ws.cell(row_num, 3).value  # Coluna C - LOTE
            historico = ws.cell(row_num, 4).value  # Coluna D - HISTÓRICO
            debito = self.parse_valor(ws.cell(row_num, 10).value)  # Coluna J - DÉBITO
            credito = self.parse_valor(ws.cell(row_num, 11).value)  # Coluna K - CRÉDITO
            saldo_atual = ws.cell(row_num, 12).value  # Coluna L - SALDO ATUAL
            
            # Verificar intervalo de datas
            if not self.esta_no_intervalo(data):
                continue
            
            lancamento = {
                'linha': row_num,
                'data': data,
                'lote': str(lote) if lote else '',
                'historico': str(historico) if historico else '',
                'debito': debito,
                'credito': credito,
                'saldo_atual': str(saldo_atual) if saldo_atual else ''
            }
            
            self.lancamentos.append(lancamento)
            total_processados += 1
            
            # Processar NFs - incluir TODAS as linhas do lote 008820001
            if historico and lote and str(lote).startswith('008820001'):
                self.processar_nf(lancamento)
            
            # Processar Recebimentos (008850001 ou RECEBIM)
            if (lote and str(lote).startswith('008850001')) or \
               (historico and 'RECEBIM' in str(historico).upper()):
                self.processar_recebimento(lancamento)
        
        print(f"✓ {total_processados} lançamentos processados")
        print(f"✓ {len(self.nfs)} NFs identificadas")
        print(f"✓ {len(self.recebimentos)} Recebimentos identificados")
    
    def processar_nf(self, lancamento):
        """Processa lançamento de NF"""
        # Extrair número da NF do histórico
        # Padrões: "NFS: 000002902", "NF: 000002902", "REF IRF NF: 000002902",
        #          "NF REC ISS:000002902 CAERN"
        historico = lancamento['historico']
        nf_match = re.search(r'NF[SE:]?[\s:]+(\d+)', historico, re.IGNORECASE)
        if not nf_match:
            nf_match = re.search(r'ISS[:\s]+(\d+)', historico, re.IGNORECASE)
        if not nf_match:
            nf_match = re.search(r'(\d{9})', historico)
        if not nf_match:
            nf_match = re.search(r'(\d{6,})', historico)

        if nf_match:
            numero_nf = nf_match.group(1)
            
            if numero_nf not in self.nfs:
                self.nfs[numero_nf] = {
                    'numero': numero_nf,
                    'total_debito': Decimal('0'),
                    'total_credito': Decimal('0'),
                    'valor_liquido': Decimal('0'),
                    'lancamentos': []
                }
            
            nf = self.nfs[numero_nf]
            nf['total_debito'] += lancamento['debito']
            nf['total_credito'] += lancamento['credito']
            nf['valor_liquido'] = nf['total_debito'] - nf['total_credito']
            nf['lancamentos'].append(lancamento)
    
    def processar_recebimento(self, lancamento):
        """Processa lançamento de Recebimento"""
        # Extrair número do recebimento
        # Padrões: "REF.RECEBIM.CR: 3  000002900" (9 dígitos)
        #          "REF.RECEBIM.CR: ND1000670"     (7 dígitos)
        historico = lancamento['historico']
        rec_match = re.search(r'RECEBIM.*?[:\s]+(?:ND)?(\d{6,})', historico, re.IGNORECASE)
        if not rec_match:
            rec_match = re.search(r'ND(\d{6,})', historico, re.IGNORECASE)
        if not rec_match:
            rec_match = re.search(r'(\d{7,})', historico)
        if not rec_match:
            rec_match = re.search(r'(\d{6,})', historico)

        if rec_match:
            numero_rec = rec_match.group(1)
            
            if numero_rec not in self.recebimentos:
                self.recebimentos[numero_rec] = {
                    'numero': numero_rec,
                    'total_debito': Decimal('0'),
                    'total_credito': Decimal('0'),
                    'valor_liquido': Decimal('0'),
                    'lancamentos': []
                }
            
            rec = self.recebimentos[numero_rec]
            rec['total_debito'] += lancamento['debito']
            rec['total_credito'] += lancamento['credito']
            rec['valor_liquido'] = rec['total_credito'] - rec['total_debito']
            rec['lancamentos'].append(lancamento)
    
    def realizar_matching(self):
        """Realiza matching entre NFs e Recebimentos"""
        nfs_matcheadas = set()
        recs_matcheados = set()
        
        # Para cada NF, procurar recebimento com valor igual
        for num_nf, nf in self.nfs.items():
            for num_rec, rec in self.recebimentos.items():
                # Comparar valores (com tolerância de 0.01)
                if abs(nf['valor_liquido'] - rec['valor_liquido']) < Decimal('0.01'):
                    self.matches.append({
                        'nf_numero': num_nf,
                        'nf_valor': nf['valor_liquido'],
                        'rec_numero': num_rec,
                        'rec_valor': rec['valor_liquido'],
                        'diferenca': nf['valor_liquido'] - rec['valor_liquido']
                    })
                    nfs_matcheadas.add(num_nf)
                    recs_matcheados.add(num_rec)
        
        # NFs não encontradas
        for num_nf, nf in self.nfs.items():
            if num_nf not in nfs_matcheadas:
                self.nao_encontrados_nf.append(nf)
        
        # Recebimentos não encontrados
        for num_rec, rec in self.recebimentos.items():
            if num_rec not in recs_matcheados:
                self.nao_encontrados_rec.append(rec)
        
        print(f"✓ {len(self.matches)} matches encontrados")
        print(f"✓ {len(self.nao_encontrados_nf)} NFs não matcheadas")
        print(f"✓ {len(self.nao_encontrados_rec)} Recebimentos não matcheados")
    
    def gerar_relatorio(self, caminho_original):
        """Gera planilha de relatório"""
        self.workbook_output = openpyxl.Workbook()
        
        # Remover aba padrão
        if 'Sheet' in self.workbook_output.sheetnames:
            del self.workbook_output['Sheet']
        
        # Criar abas
        self.criar_aba_resumo()
        self.criar_aba_nfs_detalhadas()
        self.criar_aba_recebimentos_detalhados()
        
        # Salvar arquivo
        nome_base = os.path.splitext(caminho_original)[0]
        nome_saida = f"{nome_base}_CONCILIACAO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        self.workbook_output.save(nome_saida)
        
        return nome_saida
    
    def criar_aba_resumo(self):
        """Cria aba de Resumo"""
        ws = self.workbook_output.create_sheet("1-Resumo", 0)
        
        # Calcular totais
        total_nfs = sum(nf['valor_liquido'] for nf in self.nfs.values())
        total_recebimentos = sum(rec['valor_liquido'] for rec in self.recebimentos.values())
        saldo_final = self.saldo_inicial + total_nfs - total_recebimentos
        
        # Estilo de cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Título
        ws['A1'] = "RESUMO DA CONCILIAÇÃO CONTÁBIL"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        # Período
        ws['A3'] = "Período:"
        ws['B3'] = f"{self.data_inicio.strftime('%d/%m/%Y')} a {self.data_fim.strftime('%d/%m/%Y')}"
        ws['A3'].font = Font(bold=True)
        
        # Cabeçalhos
        row = 5
        ws.cell(row, 1, "DESCRIÇÃO")
        ws.cell(row, 2, "VALOR (R$)")
        ws.cell(row, 3, "OBSERVAÇÃO")
        
        for col in range(1, 4):
            cell = ws.cell(row, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        row += 1
        self.adicionar_linha_resumo(ws, row, "Saldo Inicial", self.saldo_inicial, "Informado pelo usuário")
        
        row += 1
        self.adicionar_linha_resumo(ws, row, "Total NFs (Valor Líquido)", total_nfs, 
                                    f"{len(self.nfs)} NFs processadas")
        
        row += 1
        self.adicionar_linha_resumo(ws, row, "Total Recebimentos (Valor Líquido)", total_recebimentos,
                                    f"{len(self.recebimentos)} recebimentos processados")
        
        row += 1
        ws.cell(row, 1, "Saldo Final")
        ws.cell(row, 2, float(saldo_final))
        ws.cell(row, 2).number_format = '#,##0.00'
        ws.cell(row, 3, "Saldo Inicial + NFs - Recebimentos")
        ws.cell(row, 1).font = Font(bold=True, color="366092")
        ws.cell(row, 2).font = Font(bold=True, color="366092")
        ws.cell(row, 2).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Estatísticas de matching
        row += 3
        ws.cell(row, 1, "ESTATÍSTICAS DE MATCHING")
        ws.cell(row, 1).font = Font(bold=True, size=12, color="366092")
        
        row += 1
        self.adicionar_linha_resumo(ws, row, "Matches Encontrados", len(self.matches), 
                                    "NFs com recebimentos correspondentes")
        
        row += 1
        self.adicionar_linha_resumo(ws, row, "NFs Não Matcheadas", len(self.nao_encontrados_nf),
                                    "NFs sem recebimento correspondente")
        
        row += 1
        self.adicionar_linha_resumo(ws, row, "Recebimentos Não Matcheados", len(self.nao_encontrados_rec),
                                    "Recebimentos sem NF correspondente")
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def adicionar_linha_resumo(self, ws, row, descricao, valor, obs=""):
        """Adiciona linha na aba de resumo"""
        ws.cell(row, 1, descricao)
        
        if isinstance(valor, (Decimal, float, int)):
            ws.cell(row, 2, float(valor))
            ws.cell(row, 2).number_format = '#,##0.00'
        else:
            ws.cell(row, 2, valor)
        
        ws.cell(row, 3, obs)
        ws.cell(row, 1).font = Font(bold=True)
    
    def criar_aba_matches(self):
        """Cria aba de Matches"""
        ws = self.workbook_output.create_sheet("2-Matches")
        
        # Cabeçalho
        headers = ["NF Número", "NF Valor Líquido", "Recebimento Número", "Recebimento Valor", "Diferença"]
        self.adicionar_cabecalho(ws, headers)
        
        # Dados
        row = 2
        for match in self.matches:
            ws.cell(row, 1, match['nf_numero'])
            ws.cell(row, 2, float(match['nf_valor']))
            ws.cell(row, 2).number_format = '#,##0.00'
            ws.cell(row, 3, match['rec_numero'])
            ws.cell(row, 4, float(match['rec_valor']))
            ws.cell(row, 4).number_format = '#,##0.00'
            ws.cell(row, 5, float(match['diferenca']))
            ws.cell(row, 5).number_format = '#,##0.00'
            row += 1
        
        # Ajustar larguras
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def criar_aba_nao_encontrados(self):
        """Cria aba de Não Encontrados"""
        ws = self.workbook_output.create_sheet("3-Não Encontrados")
        
        # Seção NFs
        ws['A1'] = "NFS NÃO MATCHEADAS"
        ws['A1'].font = Font(bold=True, size=12, color="C00000")
        
        headers_nf = ["NF Número", "Total Débito", "Total Crédito", "Valor Líquido"]
        self.adicionar_cabecalho(ws, headers_nf, row_start=2)
        
        row = 3
        for nf in self.nao_encontrados_nf:
            ws.cell(row, 1, nf['numero'])
            ws.cell(row, 2, float(nf['total_debito']))
            ws.cell(row, 2).number_format = '#,##0.00'
            ws.cell(row, 3, float(nf['total_credito']))
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4, float(nf['valor_liquido']))
            ws.cell(row, 4).number_format = '#,##0.00'
            row += 1
        
        # Seção Recebimentos
        row += 2
        ws.cell(row, 1, "RECEBIMENTOS NÃO MATCHEADOS")
        ws.cell(row, 1).font = Font(bold=True, size=12, color="C00000")
        
        row += 1
        headers_rec = ["Recebimento Número", "Total Débito", "Total Crédito", "Valor Líquido"]
        self.adicionar_cabecalho(ws, headers_rec, row_start=row)
        
        row += 1
        for rec in self.nao_encontrados_rec:
            ws.cell(row, 1, rec['numero'])
            ws.cell(row, 2, float(rec['total_debito']))
            ws.cell(row, 2).number_format = '#,##0.00'
            ws.cell(row, 3, float(rec['total_credito']))
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4, float(rec['valor_liquido']))
            ws.cell(row, 4).number_format = '#,##0.00'
            row += 1
        
        # Ajustar larguras
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def criar_aba_nfs_detalhadas(self):
        """Cria aba detalhada de NFs"""
        ws = self.workbook_output.create_sheet("2-NFs Detalhadas")
        
        headers = ["NF Número", "Total Débito", "Total Crédito", "Valor Líquido", 
                   "Qtd Lançamentos", "Status"]
        self.adicionar_cabecalho(ws, headers)
        
        row = 2
        for num_nf, nf in sorted(self.nfs.items()):
            # Verificar se está matcheada
            status = "Matcheada" if any(m['nf_numero'] == num_nf for m in self.matches) else "Não Matcheada"
            
            ws.cell(row, 1, nf['numero'])
            ws.cell(row, 2, float(nf['total_debito']))
            ws.cell(row, 2).number_format = '#,##0.00'
            ws.cell(row, 3, float(nf['total_credito']))
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4, float(nf['valor_liquido']))
            ws.cell(row, 4).number_format = '#,##0.00'
            ws.cell(row, 5, len(nf['lancamentos']))
            ws.cell(row, 6, status)
            
            # Colorir linha baseado no status
            if status == "Matcheada":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for col in range(1, 7):
                ws.cell(row, col).fill = fill
            
            row += 1
        
        # Ajustar larguras
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def criar_aba_recebimentos_detalhados(self):
        """Cria aba detalhada de Recebimentos"""
        ws = self.workbook_output.create_sheet("3-Recebimentos Detalhados")
        
        headers = ["Recebimento Número", "Total Débito", "Total Crédito", "Valor Líquido",
                   "Qtd Lançamentos", "Status"]
        self.adicionar_cabecalho(ws, headers)
        
        row = 2
        for num_rec, rec in sorted(self.recebimentos.items()):
            # Verificar se está matcheado
            status = "Matcheado" if any(m['rec_numero'] == num_rec for m in self.matches) else "Não Matcheado"
            
            ws.cell(row, 1, rec['numero'])
            ws.cell(row, 2, float(rec['total_debito']))
            ws.cell(row, 2).number_format = '#,##0.00'
            ws.cell(row, 3, float(rec['total_credito']))
            ws.cell(row, 3).number_format = '#,##0.00'
            ws.cell(row, 4, float(rec['valor_liquido']))
            ws.cell(row, 4).number_format = '#,##0.00'
            ws.cell(row, 5, len(rec['lancamentos']))
            ws.cell(row, 6, status)
            
            # Colorir linha baseado no status
            if status == "Matcheado":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for col in range(1, 7):
                ws.cell(row, col).fill = fill
            
            row += 1
        
        # Ajustar larguras
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22
    
    def adicionar_cabecalho(self, ws, headers, row_start=1):
        """Adiciona cabeçalho formatado"""
        fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row_start, col, header)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
    
    def exibir_resumo(self):
        """Exibe resumo no console"""
        total_nfs = sum(nf['valor_liquido'] for nf in self.nfs.values())
        total_recebimentos = sum(rec['valor_liquido'] for rec in self.recebimentos.values())
        saldo_final = self.saldo_inicial + total_nfs - total_recebimentos
        
        print("\n" + "="*60)
        print("RESUMO DA CONCILIAÇÃO")
        print("="*60)
        print(f"Saldo Inicial:           R$ {self.saldo_inicial:,.2f}")
        print(f"Total NFs:               R$ {total_nfs:,.2f}")
        print(f"Total Recebimentos:      R$ {total_recebimentos:,.2f}")
        print(f"Saldo Final:             R$ {saldo_final:,.2f}")
        print("-"*60)
        print(f"Matches encontrados:     {len(self.matches)}")
        print(f"NFs não matcheadas:      {len(self.nao_encontrados_nf)}")
        print(f"Rec. não matcheados:     {len(self.nao_encontrados_rec)}")
        print("="*60)


def confrontar_titulos(sistema, arquivo_financeiro_path: str, saldo_inicial_float: float):
    """
    Confronta NFs não matcheadas com a planilha financeira (2-Titulos a receber).
    Coluna B = referência das NFs.
    Saldo financeiro = soma coluna K + soma coluna L.
    """
    wb_fin = openpyxl.load_workbook(arquivo_financeiro_path, data_only=True)

    # Encontrar aba "2-Titulos a receber"
    ws = None
    for sheet_name in wb_fin.sheetnames:
        if 'Titulos' in sheet_name or 'titulos' in sheet_name or 'TITULOS' in sheet_name or '2-' in sheet_name:
            ws = wb_fin[sheet_name]
            break

    if ws is None:
        abas = ', '.join(wb_fin.sheetnames)
        raise Exception(
            f"Aba de Títulos a Receber não encontrada!\n\n"
            f"O sistema procura uma aba com 'Titulos' ou '2-' no nome.\n"
            f"Abas encontradas no arquivo: {abas}\n\n"
            f"Verifique se você selecionou o arquivo correto no campo 'Planilha Financeira'."
        )

    # Ler dados da planilha financeira
    # Header na linha 1 ou 2, dados a partir da próxima
    titulos = []
    refs_financeiro = set()  # referências na coluna B (normalizadas)
    soma_col_k = Decimal('0')
    soma_col_l = Decimal('0')

    for row_num in range(2, ws.max_row + 1):
        col_b = ws.cell(row_num, 2).value  # Coluna B - referência
        col_k = ws.cell(row_num, 11).value  # Coluna K
        col_l = ws.cell(row_num, 12).value  # Coluna L

        if col_b is not None:
            ref_str = str(col_b).strip()
            if ref_str:
                titulos.append(ref_str)
                # Extrair números da referência para matching
                numeros = re.findall(r'\d+', ref_str)
                for n in numeros:
                    refs_financeiro.add(n)
                refs_financeiro.add(ref_str.upper())

        # Somar colunas K e L
        if col_k is not None:
            soma_col_k += sistema.parse_valor(col_k)
        if col_l is not None:
            soma_col_l += sistema.parse_valor(col_l)

    saldo_financeiro = soma_col_k + soma_col_l

    # Pegar NFs não matcheadas
    nfs_nao_matcheadas = sistema.nao_encontrados_nf
    encontrados = []
    nao_encontrados = []

    for nf in nfs_nao_matcheadas:
        numero_nf = nf['numero']
        # Verificar se a NF aparece nas referências do financeiro
        achou = False
        referencia_encontrada = ''

        # Tentar match direto
        if numero_nf in refs_financeiro:
            achou = True
            referencia_encontrada = numero_nf
        else:
            # Tentar match parcial (número da NF contido em alguma referência)
            for ref in titulos:
                if numero_nf in str(ref):
                    achou = True
                    referencia_encontrada = ref
                    break
                # Tentar com número sem zeros à esquerda
                nf_sem_zeros = numero_nf.lstrip('0')
                if nf_sem_zeros and nf_sem_zeros in str(ref):
                    achou = True
                    referencia_encontrada = ref
                    break

        item = {
            'nf_numero': numero_nf,
            'valor_liquido': float(nf['valor_liquido']),
            'status': 'ENCONTRADO' if achou else 'NÃO ENCONTRADO',
            'referencia_financeiro': referencia_encontrada if achou else '',
        }

        if achou:
            encontrados.append(item)
        else:
            nao_encontrados.append(item)

    return {
        'saldo_financeiro': float(saldo_financeiro),
        'soma_col_k': float(soma_col_k),
        'soma_col_l': float(soma_col_l),
        'qtd_titulos': len(titulos),
        'encontrados': encontrados,
        'nao_encontrados': nao_encontrados,
        'total_nao_matcheadas': len(nfs_nao_matcheadas),
    }


def criar_aba_confronto(wb, confronto_data):
    """Cria aba de confronto com títulos a receber no workbook de saída."""
    ws = wb.create_sheet("4-Confronto Financeiro")

    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    nao_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Título
    ws['A1'] = "CONFRONTO: NFs Não Matcheadas x Títulos a Receber"
    ws['A1'].font = Font(bold=True, size=13, color="0D9488")
    ws.merge_cells('A1:D1')

    # Resumo de saldos
    ws['A3'] = "Saldo Financeiro (col K + col L):"
    ws['B3'] = confronto_data['saldo_financeiro']
    ws['B3'].number_format = '#,##0.00'
    ws['A3'].font = Font(bold=True)

    ws['A4'] = "Soma Coluna K (Títulos Vencidos):"
    ws['B4'] = confronto_data['soma_col_k']
    ws['B4'].number_format = '#,##0.00'

    ws['A5'] = "Soma Coluna L (Títulos a Vencer):"
    ws['B5'] = confronto_data['soma_col_l']
    ws['B5'].number_format = '#,##0.00'

    # Cabeçalho da tabela
    row = 7
    headers = ["NF Número", "Valor Líquido", "Status", "Referência no Financeiro"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados - Encontrados primeiro
    row += 1
    for item in confronto_data['encontrados']:
        ws.cell(row, 1, item['nf_numero'])
        ws.cell(row, 2, item['valor_liquido'])
        ws.cell(row, 2).number_format = '#,##0.00'
        ws.cell(row, 3, item['status'])
        ws.cell(row, 4, item['referencia_financeiro'])
        for col in range(1, 5):
            ws.cell(row, col).fill = ok_fill
        row += 1

    # Não encontrados
    for item in confronto_data['nao_encontrados']:
        ws.cell(row, 1, item['nf_numero'])
        ws.cell(row, 2, item['valor_liquido'])
        ws.cell(row, 2).number_format = '#,##0.00'
        ws.cell(row, 3, item['status'])
        ws.cell(row, 4, '')
        for col in range(1, 5):
            ws.cell(row, col).fill = nao_fill
        row += 1

    # Ajustar larguras
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30


def processar_conciliacao_cliente(arquivo_path: str, saldo_inicial: str = "0",
                                   data_inicio: str = "", data_fim: str = "",
                                   output_path: str = "output.xlsx",
                                   arquivo_financeiro_path: str = None) -> dict:
    """
    Função chamada pelo app.py (web).
    Usa a classe ConciliacaoCliente sem interação de terminal.
    """
    sistema = ConciliacaoCliente()

    # Configurar parâmetros sem input()
    sistema.saldo_inicial = sistema.parse_valor(saldo_inicial)

    if data_inicio:
        try:
            sistema.data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
        except:
            sistema.data_inicio = datetime.strptime(data_inicio, "%d/%m/%Y")

    if data_fim:
        try:
            sistema.data_fim = datetime.strptime(data_fim, "%Y-%m-%d")
        except:
            sistema.data_fim = datetime.strptime(data_fim, "%d/%m/%Y")

    # Carregar e processar
    sistema.carregar_planilha(arquivo_path)
    sistema.processar_lancamentos()
    sistema.realizar_matching()

    # Gerar relatório no caminho especificado
    sistema.workbook_output = openpyxl.Workbook()
    if 'Sheet' in sistema.workbook_output.sheetnames:
        del sistema.workbook_output['Sheet']

    sistema.criar_aba_resumo()
    sistema.criar_aba_nfs_detalhadas()
    sistema.criar_aba_recebimentos_detalhados()

    # Confronto com planilha financeira (se fornecida)
    confronto_data = None
    if arquivo_financeiro_path and os.path.exists(arquivo_financeiro_path):
        saldo_inicial_float = float(sistema.saldo_inicial)
        confronto_data = confrontar_titulos(sistema, arquivo_financeiro_path, saldo_inicial_float)
        criar_aba_confronto(sistema.workbook_output, confronto_data)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    sistema.workbook_output.save(output_path)

    # Calcular totais para o frontend
    total_nfs = sum(float(nf['valor_liquido']) for nf in sistema.nfs.values())
    total_recebimentos = sum(float(rec['valor_liquido']) for rec in sistema.recebimentos.values())
    saldo_final = float(sistema.saldo_inicial) + total_nfs - total_recebimentos

    def formatar_brl(valor):
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    resultado = {
        "resumo": {
            "saldoInicial": formatar_brl(float(sistema.saldo_inicial)),
            "totalNFs": formatar_brl(total_nfs),
            "totalRecebimentos": formatar_brl(total_recebimentos),
            "saldoFinal": formatar_brl(saldo_final),
            "qtdNFs": len(sistema.nfs),
            "qtdRecebimentos": len(sistema.recebimentos),
            "qtdMatches": len(sistema.matches),
            "qtdNaoEncontrados": len(sistema.nao_encontrados_nf) + len(sistema.nao_encontrados_rec),
        }
    }

    if confronto_data:
        resultado["confronto"] = {
            "saldoInputado": formatar_brl(float(sistema.saldo_inicial)),
            "saldoFinanceiro": formatar_brl(confronto_data['saldo_financeiro']),
            "qtdTitulos": confronto_data['qtd_titulos'],
            "qtdEncontrados": len(confronto_data['encontrados']),
            "qtdNaoEncontrados": len(confronto_data['nao_encontrados']),
            "totalNaoMatcheadas": confronto_data['total_nao_matcheadas'],
        }

    return resultado


def main():
    """Função principal"""
    sistema = ConciliacaoCliente()
    sistema.iniciar()


if __name__ == "__main__":
    main()