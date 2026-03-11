# -*- coding: utf-8 -*-
"""
Pipeline completo: SA1 → API CNPJ → Confronto
──────────────────────────────────────────────
1. Lê SA1_Clientes.csv e filtra clientes dos últimos 12 meses
2. Salva clientes_12meses.xlsx (cópia de controle)
3. Para cada CNPJ consulta open.cnpja.com/office/{cnpj}
4. Extrai todas as informações disponíveis na API
5. Confronta dados da API com os dados cadastrados na SA1
6. Salva resultados incrementalmente em clientes_12meses.xlsx
7. Gera relatorio_api.xlsx com abas:
     Resumo / Dados Completos / Confronto / Divergências

Token da API (reduz rate limit):
  set CNPJA_TOKEN=seu_token_aqui   (Windows)
"""

import os
import re
import time
import random
import logging
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import httpx
import pandas as pd
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────────────────────
# CONFIGURAÇÕES
# ─────────────────────────────────────────────────────────────
PLANILHA_SA1 = Path(os.getenv(
    "PLANILHA_PATH",
    r"Z:\4 - Gestão de Receitas e Apuração de Resultados\4.4 - Núcleo de Informações\Tabelas - Protheus\SA1_Clientes.csv",
))
FALLBACK_SA1    = Path("SA1_Clientes.csv")
SAIDA_CLIENTES  = Path("clientes_12meses.xlsx")
SAIDA_RELATORIO = Path("relatorio_api.xlsx")
MESES_LIMITE    = 24
DIAS_RECONSULTA = 90   # Re-consulta CNPJs com Data Consulta mais antiga que X dias
API_BASE        = "https://open.cnpja.com/office"
API_BRASIL      = "https://brasilapi.com.br/api/cnpj/v1"
CNPJA_TOKEN     = os.getenv("CNPJA_TOKEN", "abec1069-9444-41ed-9c46-f992823f54bc-2054a7ad-1602-44f1-88bf-7d12640a47fc")
DELAY_MIN       = 8.0
DELAY_MAX       = 14.0

# Posições das colunas na SA1 (índice 0-based)
COL_CNPJ_IDX = 14   # O — CNPJ/CPF
COL_DATA_IDX = 17   # R — Último Contato

# Mapeamento: chave interna → índice da coluna SA1
COLS_SA1_IDX = {
    "nome":      2,   # C — Nome
    "fantasia":  3,   # D — Nome Reduzido
    "endereco":  5,   # F — Endereço
    "bairro":    6,   # G — Bairro
    "cep":       7,   # H — CEP
    "uf":        8,   # I — UF
    "telefone":  9,   # J — Telefone
    "municipio": 10,  # K — Município
    "email":     12,  # M — E-mail
}

# Colunas que serão adicionadas ao Excel de clientes
COLUNAS_API = [
    "Razão Social",
    "Nome Fantasia",
    "Situação",
    "Data Situação",
    "Data Abertura",
    "Porte",
    "Natureza Jurídica",
    "Capital Social",
    "CNAE Principal",
    "CNAEs Secundários",
    "Logradouro",
    "Complemento",
    "Bairro",
    "Município",
    "UF",
    "CEP",
    "Telefones",
    "Emails",
    "Simples Nacional",
    "SIMEI",
    "Sócios",
    "Matriz/Filial",
    "Data Consulta",
    "Campos Divergentes",
]
# ─────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("pipeline.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
# NORMALIZAÇÃO (para comparação sem acentos/caixa/pontuação)
# ─────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    return (
        unicodedata.normalize("NFKD", str(s))
        .encode("ascii", "ignore").decode("ascii")
        .upper().strip()
    )

# Expansão de abreviações de tipo de logradouro (primeiro token do endereço)
_ABREV_LOGRADOURO = {
    "AV.":   "AVENIDA", "AV":    "AVENIDA",
    "R.":    "RUA",     "R":     "RUA",     "RUA.":  "RUA",
    "AL.":   "ALAMEDA", "AL":    "ALAMEDA",
    "EST.":  "ESTRADA", "ESTR.": "ESTRADA",
    "ROD.":  "RODOVIA", "ROD":   "RODOVIA",
    "PR.":   "PRAIA",   "PR":    "PRAIA",
    "TV.":   "TRAVESSA","TV":    "TRAVESSA",
    "PC.":   "PRACA",   "PCA.":  "PRACA",   "PCA":   "PRACA",
    "LGO.":  "LARGO",   "LGO":   "LARGO",
    "VL.":   "VILA",
    "JD.":   "JARDIM",  "JD":    "JARDIM",
    "V.":    "VIA",
    "BLVD.": "BOULEVARD",
}

_RE_SN      = re.compile(r'\bS\.?\s*/?\s*N\.?\b')   # S/N, SN, S.N → SN
_RE_N_NUM   = re.compile(r'\bN\b\s+(\d)')            # "N 590" → "590"
_RE_PUNCT   = re.compile(r'[,./\-]')                 # pontuação → espaço
_RE_ZEROS   = re.compile(r'\b0+(\d)')                # zeros à esquerda em números
_RE_SPACES  = re.compile(r'\s+')

def _norm_logradouro(s: str) -> str:
    """Normaliza endereço: expande abreviações, remove pontuação variável e zeros."""
    base = _norm(s)
    if not base:
        return base
    words = base.split()
    words[0] = _ABREV_LOGRADOURO.get(words[0], words[0])
    base = " ".join(words)
    base = _RE_SN.sub("SN", base)
    base = _RE_N_NUM.sub(r"\1", base)
    base = _RE_PUNCT.sub(" ", base)
    base = _RE_ZEROS.sub(r"\1", base)
    return _RE_SPACES.sub(" ", base).strip()

def _norm_cep(s: str) -> str:
    return re.sub(r"\D", "", str(s))

def _norm_tel(s: str) -> str:
    digits = re.sub(r"\D", "", str(s))
    # Remove "0" inicial (formato antigo: 0 + DDD + número)
    if digits.startswith("0") and len(digits) in (10, 11):
        digits = digits[1:]
    # Remove código de país (55) + DDD para 12-13 dígitos; só DDD para 10-11
    if len(digits) in (12, 13):
        digits = digits[4:]
    elif len(digits) in (10, 11):
        digits = digits[2:]
    return digits

def limpar_cnpj(valor: str) -> str:
    return re.sub(r"\D", "", str(valor).strip())


# ─────────────────────────────────────────────────────────────
# 1. CARREGAMENTO E FILTRAGEM DA SA1
# ─────────────────────────────────────────────────────────────
def carregar_sa1() -> tuple[pd.DataFrame, str, dict]:
    path = PLANILHA_SA1
    if not path.exists() and FALLBACK_SA1.exists():
        path = FALLBACK_SA1
        log.warning(f"SA1 na rede não encontrado. Usando local: {path}")

    if not path.exists():
        raise FileNotFoundError(
            "SA1 não encontrado. Configure a variável PLANILHA_PATH "
            "ou coloque SA1_Clientes.csv na pasta do script."
        )

    log.info(f"Lendo SA1: {path}")
    ext = path.suffix.lower()
    if ext == ".csv":
        df = pd.read_csv(path, header=0, dtype=str, encoding="latin-1", sep=";")
    else:
        try:
            df = pd.read_excel(path, sheet_name="SA1_Clientes", header=0, dtype=str)
        except Exception:
            df = pd.read_excel(path, sheet_name=0, header=0, dtype=str)

    colunas = df.columns.tolist()
    def _c(idx: int) -> str:
        return colunas[idx] if idx < len(colunas) else ""

    col_cnpj = _c(COL_CNPJ_IDX)
    col_data = _c(COL_DATA_IDX)
    cols_sa1 = {chave: _c(idx) for chave, idx in COLS_SA1_IDX.items()}

    log.info(f"Coluna CNPJ: '{col_cnpj}' | Coluna Último Contato: '{col_data}'")

    df["_dt_contato"] = pd.to_datetime(df[col_data], dayfirst=True, errors="coerce")
    limite = datetime.now() - timedelta(days=MESES_LIMITE * 30)
    df = df[df["_dt_contato"] >= limite].copy()
    df = df[df[col_cnpj].notna()].copy()
    df["_cnpj_limpo"] = df[col_cnpj].apply(limpar_cnpj)

    log.info(f"Clientes filtrados (últimos {MESES_LIMITE} meses): {len(df)}")
    return df, col_cnpj, cols_sa1


# ─────────────────────────────────────────────────────────────
# 2. EXTRAÇÃO DOS DADOS DA API
# ─────────────────────────────────────────────────────────────
def extrair_dados_api(dados: dict) -> dict:
    """Extrai todos os campos do JSON retornado pela API."""
    company   = dados.get("company") or {}
    simples   = company.get("simples") or {}
    simei     = company.get("simei") or {}
    status    = dados.get("status") or {}
    address   = dados.get("address") or {}
    main_act  = dados.get("mainActivity") or {}
    side_acts = dados.get("sideActivities") or []
    phones    = dados.get("phones") or []
    emails    = dados.get("emails") or []
    members   = company.get("members") or []
    nature    = company.get("nature") or {}
    size      = company.get("size") or {}

    # Simples Nacional
    sn_since = simples.get("since")
    sn = (
        f"Optante{f' desde {sn_since}' if sn_since else ''}"
        if simples.get("optant") else "Não optante"
    )

    # SIMEI
    simei_since = simei.get("since")
    simei_txt = (
        f"Optante{f' desde {simei_since}' if simei_since else ''}"
        if simei.get("optant") else "Não optante"
    )

    # CNAEs
    cnae_principal = (
        f"{main_act.get('id', '')} - {main_act.get('text', '')}".strip(" -")
        if main_act else ""
    )
    cnaes_sec = "; ".join(
        f"{a.get('id', '')} - {a.get('text', '')}".strip(" -") for a in side_acts
    )

    # Contatos
    tel_txt   = "; ".join(f"({t.get('area', '')}){t.get('number', '')}" for t in phones)
    email_txt = "; ".join(e.get("address", "") for e in emails if e.get("address"))

    # Sócios
    socios = "; ".join(
        f"{m.get('person', {}).get('name', '')} [{m.get('role', {}).get('text', '')}]"
        for m in members
    )

    # Capital Social
    equity = company.get("equity")
    capital = (
        f"R$ {equity:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        if equity else ""
    )

    # Endereço
    rua    = address.get("street", "")
    numero = address.get("number", "")
    logradouro = f"{rua}, {numero}".strip(", ") if rua else ""

    # Matriz ou Filial
    is_head = dados.get("head")
    matriz_filial = "Matriz" if is_head is True else ("Filial" if is_head is False else "")

    # Telefone/email brutos para uso no confronto (não vão para o Excel)
    tel_bruto   = phones[0].get("area", "") + phones[0].get("number", "") if phones else ""
    email_bruto = emails[0].get("address", "") if emails else ""

    return {
        # Colunas do Excel
        "Razão Social":      company.get("name", ""),
        "Nome Fantasia":     dados.get("alias", ""),
        "Situação":          status.get("text", ""),
        "Data Situação":     dados.get("statusDate", ""),
        "Data Abertura":     dados.get("founded", ""),
        "Porte":             size.get("text", ""),
        "Natureza Jurídica": nature.get("text", ""),
        "Capital Social":    capital,
        "CNAE Principal":    cnae_principal,
        "CNAEs Secundários": cnaes_sec,
        "Logradouro":        logradouro,
        "Complemento":       address.get("details", ""),
        "Bairro":            address.get("district", ""),
        "Município":         address.get("city", ""),
        "UF":                address.get("state", ""),
        "CEP":               address.get("zip", ""),
        "Telefones":         tel_txt,
        "Emails":            email_txt,
        "Simples Nacional":  sn,
        "SIMEI":             simei_txt,
        "Sócios":            socios,
        "Matriz/Filial":     matriz_filial,
        # Campos auxiliares para confronto (prefixo _ = não gravados no Excel)
        "_tel_bruto":        tel_bruto,
        "_email_bruto":      email_bruto,
        "_logradouro_bruto": rua,
    }


def extrair_dados_brasilapi(dados: dict) -> dict:
    """Mapeia a resposta da BrasilAPI para o mesmo formato de extrair_dados_api."""
    def _tel_fmt(t: str) -> str:
        d = re.sub(r"\D", "", t)
        return f"({d[:2]}){d[2:]}" if len(d) >= 10 else t

    tipo_logr = dados.get("descricao_tipo_de_logradouro", "") or ""
    logr      = dados.get("logradouro", "") or ""
    numero    = dados.get("numero", "") or ""
    rua       = f"{tipo_logr} {logr}".strip() if tipo_logr else logr
    logradouro = f"{rua}, {numero}".strip(", ") if rua else ""

    tel1 = dados.get("ddd_telefone_1", "") or ""
    tel2 = dados.get("ddd_telefone_2", "") or ""
    tels = [_tel_fmt(t) for t in [tel1, tel2] if t.strip()]

    equity = dados.get("capital_social")
    capital = (
        f"R$ {equity:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        if equity else ""
    )

    cnae_id   = dados.get("cnae_fiscal", "")
    cnae_desc = dados.get("cnae_fiscal_descricao", "")
    cnae_principal = f"{cnae_id} - {cnae_desc}".strip(" -") if cnae_id else ""
    cnaes_sec = "; ".join(
        f"{a.get('codigo','')} - {a.get('descricao','')}".strip(" -")
        for a in (dados.get("cnaes_secundarios") or [])
    )

    socios = "; ".join(
        f"{m.get('nome_socio','')} [{m.get('qualificacao_socio','')}]"
        for m in (dados.get("qsa") or [])
    )

    simples = dados.get("opcao_pelo_simples")
    mei     = dados.get("opcao_pelo_mei")
    email   = dados.get("email", "") or ""

    return {
        "Razão Social":      dados.get("razao_social", ""),
        "Nome Fantasia":     dados.get("nome_fantasia", ""),
        "Situação":          dados.get("descricao_situacao_cadastral", ""),
        "Data Situação":     dados.get("data_situacao_cadastral", ""),
        "Data Abertura":     dados.get("data_inicio_atividade", ""),
        "Porte":             dados.get("porte", ""),
        "Natureza Jurídica": dados.get("natureza_juridica", ""),
        "Capital Social":    capital,
        "CNAE Principal":    cnae_principal,
        "CNAEs Secundários": cnaes_sec,
        "Logradouro":        logradouro,
        "Complemento":       dados.get("complemento", "") or "",
        "Bairro":            dados.get("bairro", "") or "",
        "Município":         dados.get("municipio", "") or "",
        "UF":                dados.get("uf", "") or "",
        "CEP":               dados.get("cep", "") or "",
        "Telefones":         "; ".join(tels),
        "Emails":            email,
        "Simples Nacional":  "Optante" if simples else "Não optante",
        "SIMEI":             "Optante" if mei     else "Não optante",
        "Sócios":            socios,
        "Matriz/Filial":     dados.get("descricao_identificador_matriz_filial", "") or "",
        "_tel_bruto":        re.sub(r"\D", "", tel1),
        "_email_bruto":      email,
        "_logradouro_bruto": rua,
    }


def consultar_api(client: httpx.Client, cnpj: str) -> dict | None:
    """Faz a requisição à API com retry em caso de 429."""
    resp = client.get(f"{API_BASE}/{cnpj}")

    if resp.status_code == 200:
        return extrair_dados_api(resp.json())

    if resp.status_code == 404:
        log.warning("  → Não encontrado em open.cnpja.com — tentando BrasilAPI…")
        try:
            resp_br = client.get(f"{API_BRASIL}/{cnpj}", headers={"Accept": "application/json"})
            if resp_br.status_code == 200:
                log.info("  → Encontrado via BrasilAPI")
                return extrair_dados_brasilapi(resp_br.json())
            log.warning(f"  → BrasilAPI também retornou HTTP {resp_br.status_code}")
        except Exception as e:
            log.warning(f"  → BrasilAPI falhou: {e}")
        return {"Situação": "CNPJ não encontrado"}

    if resp.status_code == 429:
        retry_after = int(resp.headers.get("Retry-After", 60))
        espera = max(retry_after, 60)
        log.warning(f"  → Rate limit. Aguardando {espera}s…")
        time.sleep(espera)
        resp = client.get(f"{API_BASE}/{cnpj}")
        if resp.status_code == 200:
            log.info("  → Retentativa OK")
            return extrair_dados_api(resp.json())
        log.warning(f"  → Retentativa falhou: HTTP {resp.status_code}")
        return {"Situação": f"Erro {resp.status_code}"}

    log.warning(f"  → Erro HTTP {resp.status_code}")
    return {"Situação": f"Erro HTTP {resp.status_code}"}


# ─────────────────────────────────────────────────────────────
# 3. CONFRONTO API × SA1
# ─────────────────────────────────────────────────────────────
def calcular_divergencias(api: dict, sa1: dict) -> str:
    """
    Compara campos da API com os cadastrados na SA1.
    Retorna string com os campos divergentes ou "✓ Sem divergências".
    """
    # ── Campos com comparação exata ──────────────────────────
    checks = {
        "Nome":      (_norm(api.get("Razão Social", "")),  _norm(sa1.get("nome", ""))),
        "CEP":       (_norm_cep(api.get("CEP", "")),       _norm_cep(sa1.get("cep", ""))),
        "Bairro":    (_norm(api.get("Bairro", "")),        _norm(sa1.get("bairro", ""))),
        "Município": (_norm(api.get("Município", "")),     _norm(sa1.get("municipio", ""))),
        "UF":        (_norm(api.get("UF", "")),            _norm(sa1.get("uf", ""))),
    }
    divergencias = [campo for campo, (v_api, v_sa1) in checks.items() if v_api != v_sa1]

    # ── Endereço: normaliza abreviações, pontuação e zeros ───
    # Usa _logradouro_bruto (pipeline em memória) ou Logradouro (do xlsx no reconfronto)
    log_api = _norm_logradouro(api.get("_logradouro_bruto") or api.get("Logradouro", ""))
    end_sa1 = _norm_logradouro(sa1.get("endereco", ""))
    if log_api and not (log_api in end_sa1 or end_sa1 in log_api):
        divergencias.append("Endereço")

    # ── Telefone: SA1 deve bater com QUALQUER telefone da API ─
    # Pula comparação se SA1 não tem telefone cadastrado
    tel_sa1 = _norm_tel(sa1.get("telefone", ""))
    if tel_sa1:
        # API pode ter vários separados por ";" (coluna Telefones) ou só o primeiro (_tel_bruto)
        tels_raw = str(api.get("Telefones", "") or api.get("_tel_bruto", "") or "")
        tels_api = [_norm_tel(t) for t in tels_raw.split(";") if t.strip()]
        tels_api = [t for t in tels_api if t]
        if tels_api and tel_sa1 not in tels_api:
            divergencias.append("Telefone")

    return ", ".join(divergencias) if divergencias else "✓ Sem divergências"


# ─────────────────────────────────────────────────────────────
# 4. RELATÓRIO FINAL
# ─────────────────────────────────────────────────────────────
def salvar_relatorio(df: pd.DataFrame, col_cnpj: str, cols_sa1: dict) -> None:
    """Gera relatorio_api.xlsx com 4 abas."""
    fill_verde    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ── Aba: Dados Completos ──────────────────────────────────
    colunas_dados = [col_cnpj] + [c for c in COLUNAS_API if c != "Campos Divergentes"]
    df_dados = df[[c for c in colunas_dados if c in df.columns]].copy()

    # ── Aba: Confronto ────────────────────────────────────────
    linhas_confronto = []
    for _, row in df.iterrows():
        if not str(row.get("Razão Social", "")).strip():
            continue

        def _sa1(key):
            col = cols_sa1.get(key, "")
            return str(row.get(col, "") or "").strip() if col else ""

        linhas_confronto.append({
            "Código":              row.get(df.columns[0], ""),
            "CNPJ":                row.get(col_cnpj, ""),
            "Razão Social (API)":  row.get("Razão Social", ""),
            "Nome (SA1)":          _sa1("nome"),
            "Nome Fantasia (API)": row.get("Nome Fantasia", ""),
            "Fantasia (SA1)":      _sa1("fantasia"),
            "CEP (API)":           row.get("CEP", ""),
            "CEP (SA1)":           _sa1("cep"),
            "Bairro (API)":        row.get("Bairro", ""),
            "Bairro (SA1)":        _sa1("bairro"),
            "Município (API)":     row.get("Município", ""),
            "Município (SA1)":     _sa1("municipio"),
            "UF (API)":            row.get("UF", ""),
            "UF (SA1)":            _sa1("uf"),
            "Logradouro (API)":    row.get("Logradouro", ""),
            "Endereço (SA1)":      _sa1("endereco"),
            "Telefone (API)":      row.get("Telefones", ""),
            "Telefone (SA1)":      _sa1("telefone"),
            "E-mail (API)":        row.get("Emails", ""),
            "E-mail (SA1)":        _sa1("email"),
            "Campos Divergentes":  row.get("Campos Divergentes", ""),
        })

    df_confronto = pd.DataFrame(linhas_confronto)

    # ── Aba: Divergências (linhas com diferença — visão completa) ─
    if not df_confronto.empty and "Campos Divergentes" in df_confronto.columns:
        df_div = df_confronto[
            ~df_confronto["Campos Divergentes"].str.startswith("✓", na=False)
        ].copy()
    else:
        df_div = pd.DataFrame()

    # ── Aba: Divergências Detalhadas (uma linha por campo divergente) ─
    # Mapeamento: nome do campo → (coluna API no df_confronto, coluna SA1)
    _MAPA_CAMPOS = {
        "Nome":      ("Razão Social (API)",  "Nome (SA1)"),
        "CEP":       ("CEP (API)",           "CEP (SA1)"),
        "Bairro":    ("Bairro (API)",        "Bairro (SA1)"),
        "Município": ("Município (API)",     "Município (SA1)"),
        "UF":        ("UF (API)",            "UF (SA1)"),
        "Endereço":  ("Logradouro (API)",    "Endereço (SA1)"),
        "Telefone":  ("Telefone (API)",      "Telefone (SA1)"),
    }
    linhas_det = []
    for _, row in df_div.iterrows():
        campos_div = [c.strip() for c in str(row.get("Campos Divergentes", "")).split(",")]
        for campo in campos_div:
            if campo not in _MAPA_CAMPOS:
                continue
            col_api, col_sa1 = _MAPA_CAMPOS[campo]
            linhas_det.append({
                "Código":       row.get("Código", ""),
                "CNPJ":         row.get("CNPJ", ""),
                "Razão Social": row.get("Razão Social (API)", ""),
                "Campo":        campo,
                "Valor (API)":  row.get(col_api, ""),
                "Valor (SA1)":  row.get(col_sa1, ""),
            })
    df_detalhes = pd.DataFrame(linhas_det) if linhas_det else pd.DataFrame(
        columns=["Código", "CNPJ", "Razão Social", "Campo", "Valor (API)", "Valor (SA1)"]
    )

    # ── Aba: Resumo ───────────────────────────────────────────
    consultados = int(
        (df["Razão Social"].notna() & (df["Razão Social"].str.strip() != "")).sum()
    ) if "Razão Social" in df.columns else 0
    erros = int(
        df["Situação"].str.startswith("Erro", na=False).sum()
    ) if "Situação" in df.columns else 0

    df_resumo = pd.DataFrame([
        {"Métrica": f"Clientes SA1 (últimos {MESES_LIMITE} meses)", "Quantidade": len(df)},
        {"Métrica": "Consultados com sucesso via API",              "Quantidade": consultados},
        {"Métrica": "Erros na consulta",                            "Quantidade": erros},
        {"Métrica": "Com divergências SA1 × API",                   "Quantidade": len(df_div)},
    ])

    with pd.ExcelWriter(SAIDA_RELATORIO, engine="openpyxl") as writer:
        df_resumo.to_excel(   writer, sheet_name="Resumo",                  index=False)
        df_dados.to_excel(    writer, sheet_name="Dados Completos",          index=False)
        df_confronto.to_excel(writer, sheet_name="Confronto",                index=False)
        df_div.to_excel(      writer, sheet_name="Divergências",             index=False)
        df_detalhes.to_excel( writer, sheet_name="Divergências Detalhadas",  index=False)

        # Coloração na aba Confronto
        if not df_confronto.empty:
            ws = writer.sheets["Confronto"]
            headers = [c.value for c in ws[1]]
            if "Campos Divergentes" in headers:
                col_idx = headers.index("Campos Divergentes")
                for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row_cells[col_idx]
                    val  = str(cell.value or "")
                    cell.fill = fill_verde if val.startswith("✓") else fill_vermelho

        # Coloração na aba Divergências (linhas todas em vermelho)
        if not df_div.empty:
            ws = writer.sheets["Divergências"]
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row_cells:
                    cell.fill = fill_vermelho

        # Coloração na aba Divergências Detalhadas
        # Coluna "Valor (API)" em laranja claro, "Valor (SA1)" em vermelho claro
        if not df_detalhes.empty:
            fill_api = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            ws = writer.sheets["Divergências Detalhadas"]
            headers = [c.value for c in ws[1]]
            idx_api = headers.index("Valor (API)") if "Valor (API)" in headers else None
            idx_sa1 = headers.index("Valor (SA1)") if "Valor (SA1)" in headers else None
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if idx_api is not None:
                    row_cells[idx_api].fill = fill_api
                if idx_sa1 is not None:
                    row_cells[idx_sa1].fill = fill_vermelho

    log.info(f"Relatório salvo → {SAIDA_RELATORIO}")
    log.info(
        f"  Resumo: {consultados} consultados | {erros} erros | "
        f"{len(df_div)} divergências"
    )


# ─────────────────────────────────────────────────────────────
# 5. MAIN
# ─────────────────────────────────────────────────────────────
def main():
    # ── Passo 1: Tenta carregar SA1 (pode estar offline) ──────
    try:
        df_sa1, col_cnpj_sa1, _ = carregar_sa1()
        df_sa1["_cnpj_limpo"] = df_sa1[col_cnpj_sa1].apply(limpar_cnpj)
    except FileNotFoundError as e:
        log.warning(f"SA1 não disponível: {e}")
        df_sa1 = None

    # ── Passo 2: Carrega ou cria o xlsx de trabalho ───────────
    if SAIDA_CLIENTES.exists():
        df = pd.read_excel(SAIDA_CLIENTES, dtype=str)
        colunas = df.columns.tolist()
        def _c(idx):
            return colunas[idx] if idx < len(colunas) else ""
        col_cnpj = _c(COL_CNPJ_IDX)
        cols_sa1 = {chave: _c(idx) for chave, idx in COLS_SA1_IDX.items()}

        # Merge: adiciona clientes novos do SA1 que ainda não estão no xlsx
        if df_sa1 is not None:
            df["_cnpj_limpo"] = df[col_cnpj].apply(limpar_cnpj)
            cnpjs_existentes  = set(df["_cnpj_limpo"].tolist())
            df_novos = df_sa1[~df_sa1["_cnpj_limpo"].isin(cnpjs_existentes)].copy()
            if not df_novos.empty:
                log.info(f"Novos clientes no SA1: {len(df_novos)} — adicionando ao xlsx…")
                # Exporta apenas colunas sem prefixo "_", alinha com o xlsx existente
                cols_exp = [c for c in df_novos.columns if not c.startswith("_")]
                df_novos = df_novos[cols_exp].copy()
                for col in df.columns:
                    if col not in df_novos.columns:
                        df_novos[col] = ""
                df_novos = df_novos.reindex(
                    columns=[c for c in df.columns if c != "_cnpj_limpo"]
                )
                df = df.drop(columns=["_cnpj_limpo"], errors="ignore")
                df = pd.concat([df, df_novos], ignore_index=True)
                log.info(f"Total após merge: {len(df)} clientes")
            else:
                log.info("Nenhum cliente novo no SA1.")
                df = df.drop(columns=["_cnpj_limpo"], errors="ignore")
    else:
        if df_sa1 is None:
            log.error("SA1 não disponível e xlsx não existe. Nada a fazer.")
            return
        df       = df_sa1
        col_cnpj = col_cnpj_sa1
        _colunas = df_sa1.columns.tolist()
        cols_sa1 = {chave: _colunas[idx] if idx < len(_colunas) else ""
                    for chave, idx in COLS_SA1_IDX.items()}
        cols_exp = [c for c in df.columns if not c.startswith("_")]
        df[cols_exp].to_excel(SAIDA_CLIENTES, index=False)
        log.info(f"Clientes filtrados salvos → {SAIDA_CLIENTES} ({len(df)} linhas)")
        df = pd.read_excel(SAIDA_CLIENTES, dtype=str)

    df["_cnpj_limpo"] = df[col_cnpj].apply(limpar_cnpj)

    # Garante que todas as colunas de saída existam
    for col in COLUNAS_API:
        if col not in df.columns:
            df[col] = ""

    # ── Passo 2: Configura cliente HTTP ───────────────────────
    if CNPJA_TOKEN:
        log.info("Autenticação: CNPJA_TOKEN configurado")
    else:
        log.warning("CNPJA_TOKEN não definido — usando API pública (rate limit restrito)")

    headers = {
        "Accept":     "application/json",
        "User-Agent": "Mozilla/5.0 (compatible; consulta-interna/1.0)",
    }
    if CNPJA_TOKEN:
        headers["Authorization"] = f"Bearer {CNPJA_TOKEN}"

    # ── Passo 3: Consulta API + Confronto ─────────────────────
    pendentes = df[df["_cnpj_limpo"].str.len() >= 11].index.tolist()
    total = len(pendentes)
    log.info(f"Total de CNPJs: {total}")

    with httpx.Client(timeout=15, headers=headers) as client:
        for i, idx in enumerate(pendentes, 1):
            cnpj = df.at[idx, "_cnpj_limpo"]

            # CPF (11 dígitos) → API de CNPJ não se aplica
            if len(cnpj) == 11:
                log.info(f"[{i}/{total}] CPF — ignorado ({cnpj})")
                continue

            # Verifica se já foi consultado e se a consulta ainda está "fresca"
            campos_div = str(df.at[idx, "Campos Divergentes"] or "").strip()
            if campos_div and campos_div.lower() not in ("", "nan"):
                data_str = str(df.at[idx, "Data Consulta"] or "").strip()
                if data_str and data_str.lower() not in ("nan", ""):
                    try:
                        dias = (datetime.now() - datetime.strptime(data_str[:10], "%Y-%m-%d")).days
                        if dias < DIAS_RECONSULTA:
                            razao = str(df.at[idx, "Razão Social"] or "").strip()
                            log.info(f"[{i}/{total}] Consultado há {dias}d — OK ({cnpj}): {razao}")
                            continue
                        log.info(f"[{i}/{total}] Re-consultando (dados de {dias}d atrás): {cnpj}")
                    except ValueError:
                        log.info(f"[{i}/{total}] Data inválida — re-consultando: {cnpj}")
                else:
                    # Sem data registrada (registros anteriores a esse controle) → mantém
                    razao = str(df.at[idx, "Razão Social"] or "").strip()
                    log.info(f"[{i}/{total}] Já consultado (sem data) ({cnpj}): {razao}")
                    continue

            situacao_ant = str(df.at[idx, "Situação"] or "").strip()
            prefixo = f"[retentativa — anterior: {situacao_ant}]" if situacao_ant else ""
            log.info(f"[{i}/{total}] Consultando: {cnpj} {prefixo}".strip())
            try:
                resultado = consultar_api(client, cnpj)
                if not resultado:
                    continue

                # Grava colunas API no DataFrame
                for col, valor in resultado.items():
                    if col in df.columns:
                        df.at[idx, col] = valor

                # Calcula confronto só quando a consulta foi bem-sucedida
                hoje = datetime.now().strftime("%Y-%m-%d")
                if resultado.get("Razão Social"):
                    sa1 = {
                        chave: str(df.at[idx, col_nome] if col_nome and col_nome in df.columns else "")
                        for chave, col_nome in cols_sa1.items()
                    }
                    df.at[idx, "Campos Divergentes"] = calcular_divergencias(resultado, sa1)
                    df.at[idx, "Data Consulta"]      = hoje
                elif resultado.get("Situação") == "CNPJ não encontrado":
                    df.at[idx, "Campos Divergentes"] = "✗ Não encontrado na API"
                    df.at[idx, "Data Consulta"]      = hoje

                log.info(
                    f"  → {resultado.get('Razão Social', '')} | "
                    f"{resultado.get('Situação', '')} | "
                    f"Simples: {resultado.get('Simples Nacional', '')} | "
                    f"Divergências: {df.at[idx, 'Campos Divergentes']}"
                )

            except Exception as e:
                df.at[idx, "Situação"] = f"Erro: {e}"
                log.error(f"  → {e}")

            # Salva incrementalmente (sem colunas auxiliares _*)
            df.drop(columns=["_cnpj_limpo"], errors="ignore").to_excel(
                SAIDA_CLIENTES, index=False
            )
            df["_cnpj_limpo"] = df[col_cnpj].apply(limpar_cnpj)
            log.info(f"  Salvo → {SAIDA_CLIENTES}")

            if i < total:
                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    # ── Passo 4: Relatório final ──────────────────────────────
    salvar_relatorio(df, col_cnpj, cols_sa1)
    log.info("Pipeline concluído.")


# ─────────────────────────────────────────────────────────────
# 6. RECONFRONTO (re-executa comparação sem chamar a API)
# ─────────────────────────────────────────────────────────────
def reconfronto():
    """
    Relê clientes_12meses.xlsx já populado pelo pipeline e recalcula
    'Campos Divergentes' com a normalização de endereços atualizada.
    """
    if not SAIDA_CLIENTES.exists():
        log.error(f"Arquivo não encontrado: {SAIDA_CLIENTES}")
        log.error("Execute pipeline primeiro para gerar o arquivo.")
        return

    log.info(f"Lendo {SAIDA_CLIENTES}…")
    df = pd.read_excel(SAIDA_CLIENTES, dtype=str)

    colunas = df.columns.tolist()
    def _c(idx):
        return colunas[idx] if idx < len(colunas) else ""

    col_cnpj = _c(COL_CNPJ_IDX)
    cols_sa1 = {chave: _c(idx) for chave, idx in COLS_SA1_IDX.items()}

    total = 0
    atualizados = 0

    for idx, row in df.iterrows():
        razao = str(row.get("Razão Social", "")).strip()
        if not razao or razao == "nan":
            continue

        total += 1
        api = row.to_dict()
        sa1 = {
            chave: str(row.get(col_nome, "") if col_nome else "").strip()
            for chave, col_nome in cols_sa1.items()
        }

        novo     = calcular_divergencias(api, sa1)
        anterior = str(row.get("Campos Divergentes", "")).strip()
        df.at[idx, "Campos Divergentes"] = novo

        if novo != anterior:
            cnpj = str(row.get(col_cnpj, "")).strip()
            log.info(f"  {cnpj} | {razao[:45]}")
            log.info(f"    antes: {anterior!r}")
            log.info(f"    agora: {novo!r}")
            atualizados += 1

    log.info(f"Reconfronto: {total} registros analisados, {atualizados} atualizados.")

    df.to_excel(SAIDA_CLIENTES, index=False)
    log.info(f"Salvo → {SAIDA_CLIENTES}")

    salvar_relatorio(df, col_cnpj, cols_sa1)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--reconfronto", action="store_true",
                        help="Recalcula divergencias sem chamar a API")
    args = parser.parse_args()

    if args.reconfronto:
        reconfronto()
    else:
        main()
